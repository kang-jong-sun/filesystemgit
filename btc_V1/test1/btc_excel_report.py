"""6엔진 교차검증 결과 → Excel 보고서"""
import sys,time,os
import numpy as np,pandas as pd
import warnings;warnings.filterwarnings('ignore')
sys.stdout.reconfigure(line_buffering=True)
from openpyxl import Workbook
from openpyxl.styles import Font,PatternFill,Alignment,Border,Side
from openpyxl.utils import get_column_letter

# Import backtest engine from full_review
from btc_full_review import load,S,ENGINES,calc_ma,adx_wilder,adx_ewm,bt

OUT=r"D:\filesystem\futures\btc_V1\test1\BTC_기획서_전체재검토_6엔진교차검증.xlsx"
YEARS=[2020,2021,2022,2023,2024,2025,2026]

def main():
    t0=time.time()
    print('='*70)
    print('  Excel Report: %d files x %d engines'%(len(S),len(ENGINES)))
    print('='*70)
    data=load()
    results=[]
    cnt=0;total=len(S)*len(ENGINES)

    for s in S:
        tf=s['tf'];d=data[tf]
        cl=d['cl'];hi=d['hi'];lo=d['lo'];vo=d['vo'];yr=d['yr'];mk=d['mk']
        mf=calc_ma(cl,vo,s['mf'],s['fp']);ms=calc_ma(cl,vo,s['ms'],s['sp'])
        wu=max(s['sp']+50,300)
        row={'f':s['f'],'desc':'%s(%d)/%s(%d) %s L%dx M%d%%'%(s['mf'],s['fp'],s['ms'],s['sp'],s['tf'],s['l'],int(s['m']*100)),'init':s['i']}
        eng_results={}

        for eng in ENGINES:
            ax=adx_wilder(hi,lo,cl,s['ap']) if eng['adx']=='w' else adx_ewm(hi,lo,cl,s['ap'])
            if eng['slip']>0:
                np.random.seed(42)
                sl2=1+np.random.uniform(-eng['slip'],eng['slip'],len(cl))
                cl2=cl*sl2;hi2=hi*sl2;lo2=lo*sl2
            else:cl2=cl;hi2=hi;lo2=lo
            r=bt(cl2,hi2,lo2,yr,mk,mf,ms,ax,s['at'],s['d'],s['sl'],s['ta'],s['tp'],s['l'],s['m'],s['i'],
                 fee=eng['fee'],skip_same=eng['skip'],strict=eng['strict'],wu=wu)
            eng_results[eng['id']]=r
            cnt+=1
            if cnt%30==0:print('  %d/%d (%.0fs)'%(cnt,total,time.time()-t0))

        # Aggregate
        rets=[eng_results[e['id']]['ret'] for e in ENGINES]
        pfs=[eng_results[e['id']]['pf'] for e in ENGINES]
        mdds=[eng_results[e['id']]['mdd'] for e in ENGINES]
        row['avg_ret']=round(np.mean(rets),1)
        row['min_ret']=round(np.min(rets),1)
        row['max_ret']=round(np.max(rets),1)
        row['avg_pf']=round(np.mean(pfs),2)
        row['min_pf']=round(np.min(pfs),2)
        row['avg_mdd']=round(np.mean(mdds),1)
        row['max_mdd']=round(np.max(mdds),1)
        row['avg_tr']=round(np.mean([eng_results[e['id']]['tr'] for e in ENGINES]),0)
        row['engines']=eng_results
        # Use E1 as primary
        e1=eng_results['E1']
        row['e1']=e1
        results.append(row)

    # Rankings
    profitable=[r for r in results if r['min_ret']>0]
    by_ret=sorted(profitable,key=lambda x:x['avg_ret'],reverse=True)[:10]
    by_stable=sorted(profitable,key=lambda x:x['max_mdd'])[:10]
    discard=sorted([r for r in results if r['min_ret']<=0 or r['max_mdd']>50 or r['min_pf']<1.0],key=lambda x:x['min_ret'])[:10]

    # ============ EXCEL ============
    print('\nGenerating Excel...')
    wb=Workbook()

    # Styles
    hdr_font=Font(bold=True,size=9)
    hdr_fill=PatternFill('solid',fgColor='4472C4')
    hdr_font_w=Font(bold=True,size=9,color='FFFFFF')
    num_fmt='#,##0'
    pct_fmt='0.0%'
    border=Border(
        left=Side(style='thin'),right=Side(style='thin'),
        top=Side(style='thin'),bottom=Side(style='thin'))
    green_fill=PatternFill('solid',fgColor='C6EFCE')
    red_fill=PatternFill('solid',fgColor='FFC7CE')
    yellow_fill=PatternFill('solid',fgColor='FFEB9C')
    center=Alignment(horizontal='center',vertical='center',wrap_text=True)

    def write_ranking_sheet(ws,title,ranked,category):
        ws.title=title
        # Headers
        headers=['순위','파일명','손익률(%)','손익금액($)']
        for y in YEARS:headers.append('%d 거래'%y)
        headers.append('총거래')
        for y in YEARS:headers.append('%d MDD(%%)'%y)
        headers.append('총MDD(%)')
        for y in YEARS:headers.append('%d PF'%y)
        headers.append('총PF')
        # 6 engine columns
        for e in ENGINES:headers.append('%s 수익(%%)'%e['n'])
        headers+=['6엔진 일치사유','비고(전략)']

        for j,h in enumerate(headers):
            c=ws.cell(row=1,column=j+1,value=h)
            c.font=hdr_font_w;c.fill=hdr_fill;c.alignment=center;c.border=border

        for i,r in enumerate(ranked):
            row_n=i+2
            e1=r['e1']
            yrly=e1.get('yearly',{})

            ws.cell(row=row_n,column=1,value=i+1).alignment=center
            ws.cell(row=row_n,column=2,value=r['f'])
            ws.cell(row=row_n,column=3,value=r['avg_ret'])
            ws.cell(row=row_n,column=4,value=round(r['avg_ret']/100*r['init'],0))

            col=5
            total_tr=0
            for y in YEARS:
                yd=yrly.get(y,{})
                tr=yd.get('tr',0);total_tr+=tr
                ws.cell(row=row_n,column=col,value=tr if tr>0 else '').alignment=center
                col+=1
            ws.cell(row=row_n,column=col,value=total_tr).alignment=center;col+=1

            for y in YEARS:
                ws.cell(row=row_n,column=col,value='').alignment=center;col+=1
            ws.cell(row=row_n,column=col,value=r['avg_mdd']).alignment=center;col+=1

            for y in YEARS:
                yd=yrly.get(y,{})
                pf=yd.get('pf',0)
                ws.cell(row=row_n,column=col,value=round(pf,1) if pf>0 else '').alignment=center
                col+=1
            ws.cell(row=row_n,column=col,value=r['avg_pf']).alignment=center;col+=1

            for e in ENGINES:
                er=r['engines'].get(e['id'],{})
                ws.cell(row=row_n,column=col,value=er.get('ret',0)).alignment=center
                col+=1

            # 6엔진 일치사유
            reasons=[]
            if r['min_ret']>0:reasons.append('6엔진전수익')
            if r['max_mdd']<=30:reasons.append('MDD<=30%')
            if r['avg_pf']>=5:reasons.append('PF>=5')
            if category=='discard':
                reasons=[]
                if r['min_ret']<=0:reasons.append('손실발생')
                if r['max_mdd']>50:reasons.append('MDD>50%')
                if r['min_pf']<1:reasons.append('PF<1')
            ws.cell(row=row_n,column=col,value=', '.join(reasons));col+=1
            ws.cell(row=row_n,column=col,value=r['desc'])

            # Coloring
            for j in range(1,col+1):
                ws.cell(row=row_n,column=j).border=border
                ws.cell(row=row_n,column=j).font=Font(size=8)
            if category=='return':
                ws.cell(row=row_n,column=3).fill=green_fill
            elif category=='stable':
                # Highlight MDD column
                pass
            elif category=='discard':
                ws.cell(row=row_n,column=3).fill=red_fill

        # Column widths
        ws.column_dimensions['A'].width=5
        ws.column_dimensions['B'].width=18
        ws.column_dimensions['C'].width=10
        ws.column_dimensions['D'].width=12
        for j in range(5,col+1):
            ws.column_dimensions[get_column_letter(j)].width=9

    # Sheet 1: Return BEST 10
    ws1=wb.active
    write_ranking_sheet(ws1,'수익률 BEST10',by_ret,'return')

    # Sheet 2: Stability BEST 10
    ws2=wb.create_sheet()
    write_ranking_sheet(ws2,'안정형 BEST10',by_stable,'stable')

    # Sheet 3: Discard BEST 10
    ws3=wb.create_sheet()
    write_ranking_sheet(ws3,'폐기형 BEST10',discard,'discard')

    # Sheet 4: Full Summary (all 46)
    ws4=wb.create_sheet('전체 요약')
    sum_headers=['파일명','전략','평균수익(%)','최소수익(%)','최대수익(%)','평균PF','최소PF',
                 '평균MDD(%)','최대MDD(%)','평균거래','판정']
    for e in ENGINES:
        sum_headers+=['%s수익(%%)'%e['n'],'%sPF'%e['n'],'%sMDD(%%)'%e['n'],'%s거래'%e['n'],'%sSL'%e['n']]

    for j,h in enumerate(sum_headers):
        c=ws4.cell(row=1,column=j+1,value=h)
        c.font=hdr_font_w;c.fill=hdr_fill;c.alignment=center;c.border=border

    for i,r in enumerate(sorted(results,key=lambda x:x['avg_ret'],reverse=True)):
        rn=i+2
        ws4.cell(row=rn,column=1,value=r['f'])
        ws4.cell(row=rn,column=2,value=r['desc'][:30])
        ws4.cell(row=rn,column=3,value=r['avg_ret'])
        ws4.cell(row=rn,column=4,value=r['min_ret'])
        ws4.cell(row=rn,column=5,value=r['max_ret'])
        ws4.cell(row=rn,column=6,value=r['avg_pf'])
        ws4.cell(row=rn,column=7,value=r['min_pf'])
        ws4.cell(row=rn,column=8,value=r['avg_mdd'])
        ws4.cell(row=rn,column=9,value=r['max_mdd'])
        ws4.cell(row=rn,column=10,value=r['avg_tr'])

        # 판정
        if r['min_ret']>0 and r['max_mdd']<=30:
            ws4.cell(row=rn,column=11,value='추천')
            ws4.cell(row=rn,column=11).fill=green_fill
        elif r['min_ret']>0:
            ws4.cell(row=rn,column=11,value='주의(MDD)')
            ws4.cell(row=rn,column=11).fill=yellow_fill
        else:
            ws4.cell(row=rn,column=11,value='폐기')
            ws4.cell(row=rn,column=11).fill=red_fill

        col=12
        for e in ENGINES:
            er=r['engines'].get(e['id'],{})
            ws4.cell(row=rn,column=col,value=er.get('ret',0));col+=1
            ws4.cell(row=rn,column=col,value=er.get('pf',0));col+=1
            ws4.cell(row=rn,column=col,value=er.get('mdd',0));col+=1
            ws4.cell(row=rn,column=col,value=er.get('tr',0));col+=1
            ws4.cell(row=rn,column=col,value=er.get('sl',0));col+=1

        for j in range(1,col):
            ws4.cell(row=rn,column=j).border=border
            ws4.cell(row=rn,column=j).font=Font(size=8)
            ws4.cell(row=rn,column=j).alignment=center

    ws4.column_dimensions['A'].width=16
    ws4.column_dimensions['B'].width=28
    for j in range(3,col):
        ws4.column_dimensions[get_column_letter(j)].width=10

    # Sheet 5: 6엔진 비교 (per strategy, per engine detail)
    ws5=wb.create_sheet('엔진별 상세')
    det_headers=['파일명']
    for e in ENGINES:
        det_headers+=['[%s]수익%%'%e['n'],'[%s]PF'%e['n'],'[%s]MDD%%'%e['n'],'[%s]거래'%e['n'],'[%s]SL'%e['n'],'[%s]FL'%e['n']]
    for j,h in enumerate(det_headers):
        c=ws5.cell(row=1,column=j+1,value=h)
        c.font=hdr_font_w;c.fill=hdr_fill;c.alignment=center;c.border=border

    for i,r in enumerate(sorted(results,key=lambda x:x['avg_ret'],reverse=True)):
        rn=i+2
        ws5.cell(row=rn,column=1,value=r['f']).font=Font(size=8,bold=True)
        col=2
        for e in ENGINES:
            er=r['engines'].get(e['id'],{})
            for val in[er.get('ret',0),er.get('pf',0),er.get('mdd',0),er.get('tr',0),er.get('sl',0),er.get('fl',0)]:
                c=ws5.cell(row=rn,column=col,value=val)
                c.font=Font(size=8);c.alignment=center;c.border=border
                # Color negative returns red
                if col%6==2 and isinstance(val,(int,float)) and val<0:
                    c.fill=red_fill
                col+=1

    ws5.column_dimensions['A'].width=16
    for j in range(2,col):ws5.column_dimensions[get_column_letter(j)].width=9

    wb.save(OUT)
    print('Excel saved: %s'%OUT)
    print('Sheets: 수익률BEST10, 안정형BEST10, 폐기형BEST10, 전체요약, 엔진별상세')
    print('\nDone in %.1f min'%((time.time()-t0)/60))

if __name__=='__main__':
    main()
