"""
v28 6엔진 교차검증 결과 → Excel 보고서 생성
로그 파일에서 데이터를 파싱하고, 상세 백테스트를 실행하여 연도별 데이터 추출
"""
import numpy as np
import pandas as pd
import re
import os
import sys
import time
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from v28_backtest_engine import (
    load_5m_data, resample_ohlcv,
    calc_ma, calc_adx_wilder, calc_rsi_wilder
)
from v28_detailed_verify import detailed_backtest

BASE = r"D:\filesystem\futures\btc_V1\test4"

# ============================================================
# 1. 전략 정의 (25개)
# ============================================================
strategies = [
    {"name": "v12.3", "tf": "5m", "ft": 0, "fp": 7, "st": 0, "sp": 100, "ap": 14, "at": 30, "rl": 30, "rh": 58, "sl": -9, "ta": 8, "tp": 6, "m": 20, "lv": 10, "d": 0, "o": 0.0, "sk": 0},
    {"name": "v13.5", "tf": "5m", "ft": 0, "fp": 7, "st": 0, "sp": 100, "ap": 14, "at": 30, "rl": 30, "rh": 58, "sl": -7, "ta": 8, "tp": 6, "m": 20, "lv": 10, "d": 0, "o": 0.0, "sk": 0},
    {"name": "v14.2", "tf": "30m", "ft": 3, "fp": 7, "st": 0, "sp": 200, "ap": 20, "at": 25, "rl": 25, "rh": 65, "sl": -7, "ta": 10, "tp": 1, "m": 30, "lv": 10, "d": 3, "o": 0.0, "sk": 0},
    {"name": "v14.4", "tf": "30m", "ft": 0, "fp": 3, "st": 0, "sp": 200, "ap": 14, "at": 35, "rl": 30, "rh": 65, "sl": -7, "ta": 6, "tp": 3, "m": 25, "lv": 10, "d": 0, "o": 0.0, "sk": 0},
    {"name": "v15.2", "tf": "30m", "ft": 0, "fp": 3, "st": 0, "sp": 200, "ap": 14, "at": 35, "rl": 30, "rh": 65, "sl": -5, "ta": 6, "tp": 5, "m": 30, "lv": 10, "d": 6, "o": 0.0, "sk": 0},
    {"name": "v15.4", "tf": "30m", "ft": 0, "fp": 3, "st": 0, "sp": 200, "ap": 14, "at": 35, "rl": 30, "rh": 65, "sl": -7, "ta": 6, "tp": 3, "m": 40, "lv": 10, "d": 0, "o": 0.0, "sk": 0},
    {"name": "v15.5", "tf": "30m", "ft": 0, "fp": 3, "st": 0, "sp": 200, "ap": 14, "at": 35, "rl": 35, "rh": 65, "sl": -7, "ta": 6, "tp": 5, "m": 35, "lv": 10, "d": 0, "o": 0.0, "sk": 0},
    {"name": "v16.0", "tf": "30m", "ft": 1, "fp": 3, "st": 0, "sp": 200, "ap": 20, "at": 35, "rl": 35, "rh": 65, "sl": -8, "ta": 4, "tp": 3, "m": 50, "lv": 10, "d": 0, "o": 0.0, "sk": 0},
    {"name": "v16.4", "tf": "30m", "ft": 1, "fp": 3, "st": 0, "sp": 200, "ap": 20, "at": 35, "rl": 35, "rh": 65, "sl": -8, "ta": 4, "tp": 3, "m": 30, "lv": 10, "d": 0, "o": 0.0, "sk": 0},
    {"name": "v16.6", "tf": "30m", "ft": 1, "fp": 3, "st": 0, "sp": 200, "ap": 20, "at": 35, "rl": 30, "rh": 70, "sl": -8, "ta": 3, "tp": 2, "m": 50, "lv": 10, "d": 5, "o": 0.0, "sk": 1},
    {"name": "v22.0F", "tf": "30m", "ft": 1, "fp": 3, "st": 0, "sp": 200, "ap": 20, "at": 35, "rl": 30, "rh": 70, "sl": -8, "ta": 3, "tp": 2, "m": 50, "lv": 10, "d": 5, "o": 0.0, "sk": 1},
    {"name": "v22.2", "tf": "30m", "ft": 0, "fp": 3, "st": 0, "sp": 200, "ap": 14, "at": 35, "rl": 30, "rh": 65, "sl": -7, "ta": 6, "tp": 3, "m": 60, "lv": 10, "d": 0, "o": 0.0, "sk": 1},
    {"name": "v22.3", "tf": "30m", "ft": 0, "fp": 3, "st": 1, "sp": 250, "ap": 20, "at": 25, "rl": 35, "rh": 65, "sl": -8, "ta": 5, "tp": 4, "m": 60, "lv": 10, "d": 0, "o": 0.0, "sk": 0},
    {"name": "v22.8", "tf": "30m", "ft": 0, "fp": 100, "st": 0, "sp": 600, "ap": 20, "at": 30, "rl": 35, "rh": 75, "sl": -8, "ta": 6, "tp": 5, "m": 35, "lv": 10, "d": 0, "o": 0.0, "sk": 1},
    {"name": "v23.4", "tf": "30m", "ft": 0, "fp": 3, "st": 0, "sp": 200, "ap": 14, "at": 35, "rl": 30, "rh": 65, "sl": -7, "ta": 6, "tp": 3, "m": 30, "lv": 10, "d": 0, "o": 0.0, "sk": 1},
    {"name": "v23.5", "tf": "10m", "ft": 0, "fp": 3, "st": 2, "sp": 200, "ap": 14, "at": 35, "rl": 40, "rh": 75, "sl": -10, "ta": 8, "tp": 4, "m": 25, "lv": 3, "d": 5, "o": 0.0, "sk": 0},
    {"name": "v23.5b", "tf": "30m", "ft": 3, "fp": 5, "st": 0, "sp": 150, "ap": 20, "at": 25, "rl": 30, "rh": 65, "sl": -10, "ta": 10, "tp": 1, "m": 25, "lv": 10, "d": 3, "o": 0.0, "sk": 0},
    {"name": "v24.2", "tf": "1h", "ft": 0, "fp": 3, "st": 0, "sp": 100, "ap": 20, "at": 30, "rl": 30, "rh": 70, "sl": -8, "ta": 6, "tp": 5, "m": 70, "lv": 10, "d": 0, "o": 0.0, "sk": 0},
    {"name": "v25.0", "tf": "5m", "ft": 0, "fp": 5, "st": 0, "sp": 100, "ap": 14, "at": 30, "rl": 40, "rh": 60, "sl": -4, "ta": 5, "tp": 3, "m": 30, "lv": 10, "d": 0, "o": 0.0, "sk": 0},
    {"name": "v25.1A", "tf": "10m", "ft": 3, "fp": 21, "st": 0, "sp": 250, "ap": 20, "at": 35, "rl": 40, "rh": 75, "sl": -6, "ta": 7, "tp": 3, "m": 50, "lv": 10, "d": 0, "o": 0.0, "sk": 0},
    {"name": "v28_T1", "tf": "15m", "ft": 1, "fp": 5, "st": 4, "sp": 300, "ap": 20, "at": 35, "rl": 35, "rh": 75, "sl": -8, "ta": 10, "tp": 5, "m": 15, "lv": 15, "d": 2, "o": 0.0, "sk": 0},
    {"name": "v28_T2", "tf": "15m", "ft": 3, "fp": 14, "st": 4, "sp": 300, "ap": 20, "at": 35, "rl": 35, "rh": 70, "sl": -7, "ta": 10, "tp": 3, "m": 50, "lv": 10, "d": 3, "o": -1.5, "sk": 0},
    {"name": "v28_T3", "tf": "15m", "ft": 4, "fp": 2, "st": 4, "sp": 300, "ap": 14, "at": 45, "rl": 30, "rh": 75, "sl": -9, "ta": 5, "tp": 5, "m": 40, "lv": 10, "d": 0, "o": 0.0, "sk": 0},
    {"name": "v32.2", "tf": "30m", "ft": 0, "fp": 100, "st": 0, "sp": 600, "ap": 20, "at": 30, "rl": 40, "rh": 80, "sl": -3, "ta": 12, "tp": 9, "m": 35, "lv": 10, "d": 0, "o": 0.0, "sk": 1},
    {"name": "v32.3", "tf": "30m", "ft": 0, "fp": 75, "st": 2, "sp": 750, "ap": 20, "at": 30, "rl": 40, "rh": 80, "sl": -3, "ta": 12, "tp": 9, "m": 35, "lv": 10, "d": 0, "o": 0.0, "sk": 1},
]

# ============================================================
# 2. 6엔진 결과 파싱 (로그에서)
# ============================================================
def parse_engine_results(log_path):
    """로그 파일에서 6엔진 결과 파싱"""
    results = {}
    current_strategy = None

    with open(log_path, 'r', encoding='utf-8', errors='ignore') as f:
        for line in f:
            # 전략 시작
            m = re.search(r'\[\d+/25\]\s+(\S+)\s+\((\S+)\)', line)
            if m:
                current_strategy = m.group(1)
                results[current_strategy] = {}

            # 엔진 결과
            m = re.search(r'E(\d)\((\S+)\s*\):\s+\$\s*([\d,]+)\s+\|\s+PF=\s*([\d.]+)\s+\|\s+MDD=\s*([\d.]+)%\s+\|\s+T=\s*(\d+)\s+W=\s*(\d+)\s+L=\s*(\d+)\s+\|\s+SL=\s*(\d+)\s+TSL=\s*(\d+)\s+REV=\s*(\d+)', line)
            if m and current_strategy:
                eng_num = int(m.group(1))
                results[current_strategy][eng_num] = {
                    'engine': m.group(2).strip(),
                    'balance': int(m.group(3).replace(',', '')),
                    'pf': float(m.group(4)),
                    'mdd': float(m.group(5)),
                    'trades': int(m.group(6)),
                    'wins': int(m.group(7)),
                    'losses': int(m.group(8)),
                    'sl': int(m.group(9)),
                    'tsl': int(m.group(10)),
                    'rev': int(m.group(11)),
                }
    return results


# ============================================================
# 3. 연도별 상세 백테스트
# ============================================================
def run_yearly_detail(s, df):
    """단일 전략의 연도별 상세 데이터"""
    ts_pd = pd.to_datetime(df['timestamp'])
    close = df['close'].values.astype(np.float64)
    high = df['high'].values.astype(np.float64)
    low = df['low'].values.astype(np.float64)
    volume = df['volume'].values.astype(np.float64)

    fast_ma = calc_ma(close, volume, s['ft'], s['fp'])
    slow_ma = calc_ma(close, volume, s['st'], s['sp'])
    adx_arr = calc_adx_wilder(high, low, close, s['ap'])
    rsi_arr = calc_rsi_wilder(close, 14)

    result = detailed_backtest(
        ts_pd, close, high, low, volume,
        fast_ma, slow_ma, adx_arr, rsi_arr,
        float(s['at']), float(s['rl']), float(s['rh']),
        float(s['sl']), float(s['ta']), float(s['tp']),
        float(s['m']), float(s['lv']),
        int(s['d']), float(s['o']),
        0.0004, int(s['sk'])
    )

    # 연도별 통계 계산
    trades = result['trades']
    yearly = result['yearly_balances']
    yearly_stats = {}

    for year in sorted(yearly.keys()):
        y = yearly[year]
        year_trades = [t for t in trades if t['entry_time'][:4] == year]
        year_wins = [t for t in year_trades if t['pnl'] > 0]
        year_losses = [t for t in year_trades if t['pnl'] <= 0]
        year_profit = sum(t['pnl'] for t in year_wins)
        year_loss_amt = sum(abs(t['pnl']) for t in year_losses)

        # 연도별 MDD 계산
        year_peak = y['start']
        year_mdd = 0
        running_bal = y['start']
        for t in year_trades:
            running_bal = t['balance_after']
            if running_bal > year_peak:
                year_peak = running_bal
            dd = (year_peak - running_bal) / year_peak * 100 if year_peak > 0 else 0
            if dd > year_mdd:
                year_mdd = dd

        yearly_stats[year] = {
            'start': y['start'],
            'end': y['end'],
            'return_pct': (y['end'] - y['start']) / y['start'] * 100 if y['start'] > 0 else 0,
            'trades': len(year_trades),
            'wins': len(year_wins),
            'losses': len(year_losses),
            'pf': year_profit / year_loss_amt if year_loss_amt > 0 else (999 if year_profit > 0 else 0),
            'mdd': year_mdd,
        }

    return result, yearly_stats


# ============================================================
# 4. Excel 생성
# ============================================================
def create_excel(all_data, output_path):
    """Excel 보고서 생성"""
    wb = Workbook()

    # 스타일 정의
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_fill_green = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
    header_fill_red = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    header_fill_orange = PatternFill(start_color="BF8F00", end_color="BF8F00", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    profit_fill = PatternFill(start_color="E2EFDA", fill_type="solid")
    loss_fill = PatternFill(start_color="FCE4EC", fill_type="solid")
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    headers = ['순위', '파일명', '손익률', '손익금액', '2020 거래', '2021 거래', '2022 거래',
               '2023 거래', '2024 거래', '2025 거래', '2026 거래', '총 거래량',
               '2020 MDD', '2021 MDD', '2022 MDD', '2023 MDD', '2024 MDD', '2025 MDD', '2026 MDD',
               '2020 PF', '2021 PF', '2022 PF', '2023 PF', '2024 PF', '2025 PF', '2026 PF',
               'E1-E5 일치', '6엔진 편차%', '비고(사유)']

    def write_sheet(ws, title, fill_color, data_list):
        ws.title = title
        # 제목 행
        ws.merge_cells('A1:AB1')
        ws['A1'] = title
        ws['A1'].font = Font(bold=True, size=14)

        # 헤더
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = fill_color
            cell.alignment = center_align
            cell.border = thin_border

        # 데이터
        for row_idx, d in enumerate(data_list, 4):
            ys = d.get('yearly_stats', {})

            values = [
                d['rank'],
                d['name'],
                f"{d['return_pct']:+.1f}%",
                f"${d['balance']:,.0f}",
            ]

            # 연도별 거래수
            for yr in ['2020', '2021', '2022', '2023', '2024', '2025', '2026']:
                if yr in ys:
                    values.append(ys[yr]['trades'])
                else:
                    values.append(0)

            values.append(d['trades'])

            # 연도별 MDD
            for yr in ['2020', '2021', '2022', '2023', '2024', '2025', '2026']:
                if yr in ys and ys[yr]['mdd'] > 0:
                    values.append(f"{ys[yr]['mdd']:.1f}%")
                else:
                    values.append("-")

            # 연도별 PF
            for yr in ['2020', '2021', '2022', '2023', '2024', '2025', '2026']:
                if yr in ys and ys[yr]['pf'] > 0:
                    pf_val = ys[yr]['pf']
                    values.append(f"{pf_val:.2f}" if pf_val < 999 else "INF")
                else:
                    values.append("-")

            values.append("O" if d['e1_e5_match'] else "X")
            values.append(f"{d['deviation']:.1f}%")
            values.append(d['note'])

            for col_idx, val in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.alignment = center_align
                cell.border = thin_border
                # 손익률 색상
                if col_idx == 3:
                    if d['return_pct'] > 0:
                        cell.fill = profit_fill
                    else:
                        cell.fill = loss_fill

        # 열 너비 조정
        col_widths = [5, 12, 10, 12] + [7]*7 + [8] + [8]*7 + [7]*7 + [8, 8, 30]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    # ============================================================
    # 데이터 정렬
    # ============================================================
    # 수익률 BEST 10
    profitable = sorted([d for d in all_data if d['return_pct'] > 0],
                        key=lambda x: x['return_pct'], reverse=True)[:10]
    for i, d in enumerate(profitable):
        d['rank'] = i + 1

    # 안정형 BEST 10 (수익 > 0, MDD 기준)
    stable = sorted([d for d in all_data if d['return_pct'] > 0],
                    key=lambda x: x['mdd'])[:10]
    for i, d in enumerate(stable):
        d['rank'] = i + 1

    # 폐기형 BEST 10 (손실 or 높은 MDD or 낮은 PF)
    discard = sorted(all_data, key=lambda x: x['return_pct'])[:10]
    for i, d in enumerate(discard):
        d['rank'] = i + 1

    # 시트 생성
    ws1 = wb.active
    write_sheet(ws1, "수익률 BEST 10", header_fill_green, profitable)

    ws2 = wb.create_sheet()
    write_sheet(ws2, "안정형 BEST 10", header_fill, stable)

    ws3 = wb.create_sheet()
    write_sheet(ws3, "폐기형 BEST 10", header_fill_red, discard)

    # 전체 현황 시트
    ws4 = wb.create_sheet("전체 현황")
    ws4.merge_cells('A1:AB1')
    ws4['A1'] = '전체 25개 전략 6엔진 교차검증 결과 (수익률 순)'
    ws4['A1'].font = Font(bold=True, size=14)

    all_sorted = sorted(all_data, key=lambda x: x['return_pct'], reverse=True)
    for i, d in enumerate(all_sorted):
        d['rank'] = i + 1
    write_sheet_data(ws4, headers, header_fill_orange, all_sorted)

    # 6엔진 상세 시트
    ws5 = wb.create_sheet("6엔진 상세")
    ws5['A1'] = '6엔진 교차검증 상세 매트릭스'
    ws5['A1'].font = Font(bold=True, size=14)
    eng_headers = ['전략', 'E1 잔액', 'E1 PF', 'E2 잔액', 'E2 PF', 'E3 잔액', 'E3 PF',
                   'E4 잔액', 'E4 PF', 'E5 잔액', 'E5 PF', 'E6 잔액', 'E6 PF',
                   'E1-E5 편차', 'E6 편차', '판정']
    for col_idx, h in enumerate(eng_headers, 1):
        cell = ws5.cell(row=3, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    for row_idx, d in enumerate(all_sorted, 4):
        eng = d.get('engine_data', {})
        vals = [d['name']]
        for e in [1, 2, 3, 4, 5, 6]:
            if e in eng:
                vals.append(f"${eng[e]['balance']:,}")
                vals.append(f"{eng[e]['pf']:.2f}")
            else:
                vals.extend(['-', '-'])
        # E1-E5 편차
        e1_5 = [eng[e]['balance'] for e in [1, 2, 3, 4, 5] if e in eng]
        if e1_5:
            dev = (max(e1_5) - min(e1_5)) / np.mean(e1_5) * 100 if np.mean(e1_5) > 0 else 0
            vals.append(f"{dev:.2f}%")
        else:
            vals.append("-")
        # E6 편차
        if 6 in eng and 1 in eng:
            e6_dev = abs(eng[6]['balance'] - eng[1]['balance']) / eng[1]['balance'] * 100 if eng[1]['balance'] > 0 else 0
            vals.append(f"{e6_dev:.0f}%")
        else:
            vals.append("-")
        vals.append("PASS" if d['e1_e5_match'] else "FAIL")

        for col_idx, val in enumerate(vals, 1):
            cell = ws5.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = center_align
            cell.border = thin_border

    for i in range(1, 17):
        ws5.column_dimensions[get_column_letter(i)].width = 12

    wb.save(output_path)
    print(f"Excel saved: {output_path}")


def write_sheet_data(ws, headers, fill, data_list):
    """시트에 데이터 쓰기 (전체현황용)"""
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    header_font = Font(bold=True, size=11, color="FFFFFF")
    profit_fill = PatternFill(start_color="E2EFDA", fill_type="solid")
    loss_fill = PatternFill(start_color="FCE4EC", fill_type="solid")

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = fill
        cell.alignment = center_align
        cell.border = thin_border

    for row_idx, d in enumerate(data_list, 4):
        ys = d.get('yearly_stats', {})
        values = [d['rank'], d['name'], f"{d['return_pct']:+.1f}%", f"${d['balance']:,.0f}"]
        for yr in ['2020','2021','2022','2023','2024','2025','2026']:
            values.append(ys.get(yr, {}).get('trades', 0))
        values.append(d['trades'])
        for yr in ['2020','2021','2022','2023','2024','2025','2026']:
            m = ys.get(yr, {}).get('mdd', 0)
            values.append(f"{m:.1f}%" if m > 0 else "-")
        for yr in ['2020','2021','2022','2023','2024','2025','2026']:
            p = ys.get(yr, {}).get('pf', 0)
            values.append(f"{p:.2f}" if 0 < p < 999 else ("INF" if p >= 999 else "-"))
        values.append("O" if d['e1_e5_match'] else "X")
        values.append(f"{d['deviation']:.1f}%")
        values.append(d['note'])

        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = center_align
            cell.border = thin_border
            if col_idx == 3:
                cell.fill = profit_fill if d['return_pct'] > 0 else loss_fill

    col_widths = [5, 12, 10, 12] + [7]*7 + [8] + [8]*7 + [7]*7 + [8, 8, 30]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ============================================================
# 5. 메인
# ============================================================
def main():
    print("=" * 70)
    print("  Excel Report Generator: 6엔진 교차검증 결과")
    print("=" * 70)

    # 1) 로그에서 6엔진 결과 파싱
    log_path = os.path.join(BASE, "v28_full_review_log.txt")
    print(f"\n[1] Parsing engine results from log...")
    engine_results = parse_engine_results(log_path)
    print(f"  Parsed {len(engine_results)} strategies")

    # 2) 데이터 로드
    print(f"\n[2] Loading data for yearly analysis...")
    df_5m = load_5m_data(BASE)
    tf_map = {
        '5m': df_5m,
        '10m': resample_ohlcv(df_5m, 10),
        '15m': resample_ohlcv(df_5m, 15),
        '30m': resample_ohlcv(df_5m, 30),
        '1h': resample_ohlcv(df_5m, 60),
    }

    # 3) 각 전략 상세 분석
    print(f"\n[3] Running yearly detailed analysis for {len(strategies)} strategies...")
    all_data = []

    for idx, s in enumerate(strategies):
        name = s['name']
        tf = s['tf']
        df = tf_map[tf]

        print(f"  [{idx+1}/{len(strategies)}] {name} ({tf})...", end=" ")
        t0 = time.time()

        # 상세 백테스트
        detail_result, yearly_stats = run_yearly_detail(s, df)
        elapsed = time.time() - t0

        # 6엔진 데이터
        eng = engine_results.get(name, {})
        e1_bal = eng.get(1, {}).get('balance', detail_result['balance'])

        # E1-E5 일관성
        e1_5_bals = [eng[e]['balance'] for e in [1, 2, 3, 4, 5] if e in eng]
        if e1_5_bals and np.mean(e1_5_bals) > 0:
            e1_5_dev = (max(e1_5_bals) - min(e1_5_bals)) / np.mean(e1_5_bals) * 100
            e1_e5_match = e1_5_dev < 5.0
        else:
            e1_5_dev = 0
            e1_e5_match = True

        # 전체 편차 (E6 포함)
        all_bals = [eng[e]['balance'] for e in [1, 2, 3, 4, 5, 6] if e in eng]
        if all_bals and np.mean(all_bals) > 0:
            total_dev = np.std(all_bals) / np.mean(all_bals) * 100
        else:
            total_dev = 0

        # 비고 생성
        notes = []
        if detail_result['balance'] < 3000:
            notes.append("손실전략")
        if detail_result['mdd'] > 50:
            notes.append(f"고위험MDD{detail_result['mdd']:.0f}%")
        if detail_result['total_trades'] < 30:
            notes.append(f"소표본{detail_result['total_trades']}건")
        if e1_e5_match:
            notes.append("E1-E5일치")
        else:
            notes.append(f"E1-E5편차{e1_5_dev:.1f}%")
        if not eng:
            notes.append("엔진데이터없음")

        entry = {
            'name': name,
            'balance': detail_result['balance'],
            'return_pct': (detail_result['balance'] - 3000) / 3000 * 100,
            'trades': detail_result['total_trades'],
            'pf': eng.get(1, {}).get('pf', 0),
            'mdd': detail_result['mdd'],
            'sl_count': sum(1 for t in detail_result['trades'] if t['close_type'] == 'SL'),
            'tsl_count': sum(1 for t in detail_result['trades'] if t['close_type'] == 'TSL'),
            'rev_count': sum(1 for t in detail_result['trades'] if t['close_type'] == 'REV'),
            'yearly_stats': yearly_stats,
            'engine_data': eng,
            'e1_e5_match': e1_e5_match,
            'deviation': total_dev,
            'note': "; ".join(notes),
            'rank': 0,
        }
        all_data.append(entry)
        print(f"${detail_result['balance']:,.0f} | PF={entry['pf']:.2f} | MDD={detail_result['mdd']:.1f}% | {elapsed:.1f}s")

    # 4) Excel 생성
    print(f"\n[4] Generating Excel report...")
    output_path = os.path.join(BASE, "v28_전체기획서_6엔진_교차검증_결과.xlsx")
    create_excel(all_data, output_path)

    print(f"\n{'='*70}")
    print(f"  완료! {output_path}")
    print(f"{'='*70}")


if __name__ == '__main__':
    main()
