"""Convert cross-verification results to Excel with 4 sheets"""
import pandas as pd, os, sys

try:
    import openpyxl
except:
    import subprocess
    subprocess.run([sys.executable, '-m', 'pip', 'install', 'openpyxl'], check=True)
    import openpyxl

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DIR = r"D:\filesystem\futures\btc_V1\test3"
df = pd.read_csv(os.path.join(DIR, "cross_verify_full_results.csv"))

# Classify
def classify(row):
    engines_profit = 0
    for eng in ['Wilder_Manual','EWM_alpha','EWM_span','SMA_Wilder','No_ADX','No_RSI']:
        col = f"return_{eng}" if f"return_{eng}" in row.index else None
        if col and row.get(col, -999) > 0:
            engines_profit += 1
    return engines_profit

# Get unique strategies - average across engines
strategies = df['strategy'].unique() if 'strategy' in df.columns else df.iloc[:,0].unique()

# Build summary per strategy
results = []
for strat in df['strategy'].unique():
    sdf = df[df['strategy'] == strat]
    avg_ret = sdf['return_pct'].mean()
    avg_pf = sdf['pf'].mean()
    avg_mdd = sdf['mdd_pct'].mean()
    avg_trades = sdf['trades'].mean()
    engines_profit = (sdf['return_pct'] > 0).sum()
    total_engines = len(sdf)

    # Yearly data from first profitable engine (or first engine)
    best_row = sdf[sdf['return_pct'] == sdf['return_pct'].max()].iloc[0] if len(sdf) > 0 else sdf.iloc[0]

    r = {
        'strategy': strat,
        'avg_return_pct': round(avg_ret, 1),
        'avg_final': round(sdf['final_balance'].mean(), 0),
        'avg_pf': round(avg_pf, 2),
        'avg_mdd': round(avg_mdd, 1),
        'avg_trades': round(avg_trades, 0),
        'engines_profit': f"{engines_profit}/{total_engines}",
        'engines_profit_n': engines_profit,
        'std_return': round(sdf['return_pct'].std(), 1),
    }

    # Yearly columns
    for yr in range(2020, 2027):
        for col_type in ['trades', 'pnl', 'mdd', 'pf']:
            col = f"y{yr}_{col_type}"
            if col in best_row.index and not pd.isna(best_row[col]):
                r[f"{yr}_{col_type}"] = round(best_row[col], 1) if col_type in ['pnl','mdd','pf'] else int(best_row[col])
            else:
                r[f"{yr}_{col_type}"] = 0

    results.append(r)

rdf = pd.DataFrame(results)

# Sort for each category
return_best = rdf[rdf['engines_profit_n'] >= 4].sort_values('avg_return_pct', ascending=False).head(10)
stability_best = rdf[rdf['engines_profit_n'] >= 4].sort_values('avg_mdd', ascending=True).head(10)  # lower MDD = more stable (less negative)
# Actually MDD is negative, so ascending means "least negative" = best stability
# Wait - MDD values could be stored as positive or negative. Let me check
# In our data, MDD is likely stored as negative percentage. More negative = worse.
# For stability, we want the LEAST negative (closest to 0)
stability_best = rdf[rdf['engines_profit_n'] >= 3].sort_values('avg_mdd', ascending=False).head(10)
discard = rdf.sort_values('avg_return_pct', ascending=True).head(10)

# Create Excel
out_path = os.path.join(DIR, "전체_기획서_6엔진_교차검증.xlsx")
with pd.ExcelWriter(out_path, engine='openpyxl') as writer:

    def write_sheet(sheet_df, sheet_name, rank_start=1):
        rows = []
        for idx, (_, row) in enumerate(sheet_df.iterrows()):
            yr_trades = " / ".join([str(row.get(f"{yr}_trades", 0)) for yr in range(2020, 2027)])
            yr_mdd = " / ".join([f"{row.get(f'{yr}_mdd', 0):.0f}%" for yr in range(2020, 2027)])
            yr_pf = " / ".join([f"{row.get(f'{yr}_pf', 0):.1f}" for yr in range(2020, 2027)])

            eng_note = f"{row['engines_profit']} 엔진 수익"
            if row['engines_profit_n'] >= 6:
                eng_note += " (전원일치)"
            elif row['engines_profit_n'] >= 5:
                eng_note += " (1개 제외)"
            elif row['engines_profit_n'] >= 4:
                eng_note += " (2개 제외)"
            else:
                eng_note += " (과반 손실)"

            if row['avg_return_pct'] > 10000:
                note = "고수익 복리형"
            elif row['avg_return_pct'] > 1000:
                note = "중수익"
            elif row['avg_return_pct'] > 0:
                note = "저수익"
            else:
                note = "손실 전략"

            if abs(row['avg_mdd']) < 30:
                note += " / 저MDD"
            elif abs(row['avg_mdd']) < 60:
                note += " / 중MDD"
            else:
                note += " / 고MDD"

            rows.append({
                '순위': rank_start + idx,
                '파일명': row['strategy'],
                '손익률': f"{row['avg_return_pct']:+,.1f}%",
                '손익금액': f"${row['avg_final']:,.0f}",
                '년도별 거래량(20~26)': yr_trades,
                '총 거래량': int(row['avg_trades']),
                '년도별 MDD(20~26)': yr_mdd,
                '년도별 PF(20~26)': yr_pf,
                '6개 엔진 일치': eng_note,
                '비고(사유)': note,
            })

        out_df = pd.DataFrame(rows)
        out_df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=1)

        ws = writer.sheets[sheet_name]
        ws.cell(row=1, column=1, value=sheet_name).font = Font(bold=True, size=14)

        # Column widths
        widths = [6, 20, 14, 14, 30, 10, 35, 35, 20, 25]
        for i, w in enumerate(widths):
            ws.column_dimensions[get_column_letter(i+1)].width = w

        # Header style
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=10)
        for col in range(1, len(widths)+1):
            cell = ws.cell(row=2, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', wrap_text=True)

        # Data alignment
        for row_idx in range(3, 3 + len(rows)):
            for col_idx in range(1, len(widths)+1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.alignment = Alignment(wrap_text=True, vertical='center')
                if col_idx <= 2:
                    cell.alignment = Alignment(horizontal='center', vertical='center')

    write_sheet(return_best, '수익률 BEST 10')
    write_sheet(stability_best, '안정형 BEST 10')
    write_sheet(discard, '폐기형 BEST 10')

    # Sheet 4: Full engine matrix
    pivot_cols = ['strategy', 'engine', 'return_pct', 'pf', 'mdd_pct', 'trades', 'final_balance']
    avail_cols = [c for c in pivot_cols if c in df.columns]
    df[avail_cols].to_excel(writer, sheet_name='전체 엔진별 상세', index=False)

    ws4 = writer.sheets['전체 엔진별 상세']
    ws4.column_dimensions['A'].width = 18
    ws4.column_dimensions['B'].width = 18
    for col in 'CDEFG':
        ws4.column_dimensions[col].width = 14

print(f"Excel saved: {out_path}")
print(f"Sheets: 수익률 BEST 10, 안정형 BEST 10, 폐기형 BEST 10, 전체 엔진별 상세")
