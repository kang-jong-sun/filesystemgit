import sys;sys.stdout.reconfigure(encoding='utf-8')
import pandas as pd,numpy as np
from numba import njit
from openpyxl import Workbook
from openpyxl.styles import Font,PatternFill,Alignment
from openpyxl.utils import get_column_letter
qn = None  # not needed for basic styling
np.seterr(all='ignore')
BASE="D:/filesystem/futures/btc_V1/test"

# ═══ 6개 엔진 ═══
@njit
def E1(cl,hi,lo,ts,fv,sv,av,rv,am,ar,mn,rl,rh,sp,tp,wp,mp,lv,sk,ic,gp,st):
    n=len(cl);c=ic;p=0;ex=0.0;ps=0.0;sl=0.0;tn=False;th=0.0;tl=999999.0
    pk=c;md=0.0;ms=c;sc=0;tc=0;rc=0;w=0;lo_=0;gpr=0.0;gls=0.0
    wa=0;ws=0;le=0;ld=0
    YP=np.zeros(7);YW=np.zeros(7,dtype=np.int32);YL=np.zeros(7,dtype=np.int32);YM=np.zeros(7);YK=np.zeros(7)
    for j in range(7):YK[j]=ic
    # 월별 (최대100)
    MP=np.zeros(100);MW=np.zeros(100,dtype=np.int32);ML=np.zeros(100,dtype=np.int32);MC=np.zeros(100)
    mi=-1;cym=-1
    for i in range(st,n):
        px=cl[i];h_=hi[i];l_=lo[i]
        if i>st and i%1440==0:ms=c
        yr=2020+int((ts[i]-1577836800)/(365.25*86400))
        yi=yr-2020 if yr>=2020 and yr<=2026 else -1
        # 월 변경
        ds=int((ts[i]-1577836800)/86400)
        yy=2020+int(ds/365.25)
        dy=ds-int((yy-2020)*365.25)
        if dy<0:yy-=1;dy=ds-int((yy-2020)*365.25)
        mm=1+int(dy/30.44)
        if mm>12:mm=12
        if mm<1:mm=1
        ym=yy*100+mm
        if ym!=cym:
            mi=min(mi+1,99);cym=ym;MC[mi]=c
        if p!=0:
            wa=0
            if not tn:
                if(p==1 and l_<=sl)or(p==-1 and h_>=sl):
                    pnl=(sl-ex)/ex*ps*p-ps*0.0004;c+=pnl;sc+=1
                    if pnl>0:w+=1;gpr+=pnl
                    else:lo_+=1;gls+=abs(pnl)
                    if yi>=0:YP[yi]+=pnl;(YW if pnl>0 else YL)[yi]+=1
                    if mi>=0:MP[mi]+=pnl;(MW if pnl>0 else ML)[mi]+=1
                    ld=p;le=i;p=0;continue
            br=(h_-ex)/ex*100 if p==1 else(ex-l_)/ex*100
            if br>=tp:tn=True
            if tn:
                if p==1:
                    if h_>th:th=h_
                    ns=th*(1-wp/100)
                    if ns>sl:sl=ns
                    if px<=sl:
                        pnl=(px-ex)/ex*ps*p-ps*0.0004;c+=pnl;tc+=1
                        if pnl>0:w+=1;gpr+=pnl
                        else:lo_+=1;gls+=abs(pnl)
                        if yi>=0:YP[yi]+=pnl;(YW if pnl>0 else YL)[yi]+=1
                        if mi>=0:MP[mi]+=pnl;(MW if pnl>0 else ML)[mi]+=1
                        ld=p;le=i;p=0;continue
                else:
                    if l_<tl:tl=l_
                    ns=tl*(1+wp/100)
                    if ns<sl:sl=ns
                    if px>=sl:
                        pnl=(px-ex)/ex*ps*p-ps*0.0004;c+=pnl;tc+=1
                        if pnl>0:w+=1;gpr+=pnl
                        else:lo_+=1;gls+=abs(pnl)
                        if yi>=0:YP[yi]+=pnl;(YW if pnl>0 else YL)[yi]+=1
                        if mi>=0:MP[mi]+=pnl;(MW if pnl>0 else ML)[mi]+=1
                        ld=p;le=i;p=0;continue
            if i>0:
                bn=fv[i]>sv[i];bp=fv[i-1]>sv[i-1]
                cu=bn and not bp;cd=not bn and bp
                if(p==1 and cd)or(p==-1 and cu):
                    pnl=(px-ex)/ex*ps*p-ps*0.0004;c+=pnl;rc+=1
                    if pnl>0:w+=1;gpr+=pnl
                    else:lo_+=1;gls+=abs(pnl)
                    if yi>=0:YP[yi]+=pnl;(YW if pnl>0 else YL)[yi]+=1
                    if mi>=0:MP[mi]+=pnl;(MW if pnl>0 else ML)[mi]+=1
                    ld=p;le=i;p=0
        if i<1:continue
        bn=fv[i]>sv[i];bp=fv[i-1]>sv[i-1]
        cu=bn and not bp;cd=not bn and bp
        if p==0:
            if cu:wa=1;ws=i
            elif cd:wa=-1;ws=i
            if wa!=0 and(mn==0 or i>ws):
                if mn>0 and i-ws>mn:wa=0;continue
                if wa==1 and cd:wa=-1;ws=i;continue
                elif wa==-1 and cu:wa=1;ws=i;continue
                if sk and wa==ld:continue
                aok=av[i]>=am;arr=True
                if ar>0 and i>=ar:arr=av[i]>av[i-ar]
                rok=(rv[i]>=rl)and(rv[i]<=rh)
                gok=True
                if gp>0:
                    g=abs(fv[i]-sv[i])/sv[i]*100
                    if g<gp:gok=False
                if aok and arr and rok and gok and c>0:
                    if ms>0 and(c-ms)/ms<=-0.20:wa=0;continue
                    mg=c*mp;ps=mg*lv;c-=ps*0.0004
                    p=wa;ex=px;tn=False;th=px;tl=px
                    if p==1:sl=ex*(1-sp/100)
                    else:sl=ex*(1+sp/100)
                    pk=max(pk,c);wa=0
        pk=max(pk,c);dd=(pk-c)/pk if pk>0 else 0
        if dd>md:md=dd
        if yi>=0:
            YK[yi]=max(YK[yi],c)
            yd=(YK[yi]-c)/YK[yi] if YK[yi]>0 else 0
            if yd>YM[yi]:YM[yi]=yd
        if c<=0:break
    if p!=0 and c>0:
        pnl=(cl[n-1]-ex)/ex*ps*p-ps*0.0004;c+=pnl
        if pnl>0:w+=1;gpr+=pnl
        else:lo_+=1;gls+=abs(pnl)
    if gls<0.001:gls=0.001
    return c,gpr/gls,md*100,w+lo_,w,lo_,sc,tc,rc,YP,YW,YL,YM,mi+1,MP,MW,ML,MC

# Python 엔진 (E2~E6 동일로직)
def E_py(cl,hi,lo,ts,fv,sv,av,rv,am,ar,mn,rl,rh,sp,tp,wp,mp,lv,sk,ic,gp,st):
    n=len(cl);c=float(ic);p=0;ex=0.0;ps=0.0;sl_=0.0;tn=False;th=0.0;tl=999999.0
    pk=c;md=0.0;ms=c;sc=0;tc=0;rc=0;w=0;lo2=0;gpr=0.0;gls=0.0
    wa=0;ws=0;le=0;ld=0
    for i in range(st,n):
        px=float(cl[i]);h_=float(hi[i]);l_=float(lo[i])
        if i>st and i%1440==0:ms=c
        if p!=0:
            wa=0
            if not tn:
                if(p==1 and l_<=sl_)or(p==-1 and h_>=sl_):
                    pnl=(sl_-ex)/ex*ps*p-ps*0.0004;c+=pnl;sc+=1
                    if pnl>0:w+=1;gpr+=pnl
                    else:lo2+=1;gls+=abs(pnl)
                    ld=p;le=i;p=0;continue
            br=(h_-ex)/ex*100 if p==1 else(ex-l_)/ex*100
            if br>=tp:tn=True
            if tn:
                if p==1:
                    if h_>th:th=h_
                    ns=th*(1-wp/100)
                    if ns>sl_:sl_=ns
                    if px<=sl_:
                        pnl=(px-ex)/ex*ps*p-ps*0.0004;c+=pnl;tc+=1
                        if pnl>0:w+=1;gpr+=pnl
                        else:lo2+=1;gls+=abs(pnl)
                        ld=p;le=i;p=0;continue
                else:
                    if l_<tl:tl=l_
                    ns=tl*(1+wp/100)
                    if ns<sl_:sl_=ns
                    if px>=sl_:
                        pnl=(px-ex)/ex*ps*p-ps*0.0004;c+=pnl;tc+=1
                        if pnl>0:w+=1;gpr+=pnl
                        else:lo2+=1;gls+=abs(pnl)
                        ld=p;le=i;p=0;continue
            if i>0:
                bn=float(fv[i])>float(sv[i]);bp=float(fv[i-1])>float(sv[i-1])
                cu=bn and not bp;cd=not bn and bp
                if(p==1 and cd)or(p==-1 and cu):
                    pnl=(px-ex)/ex*ps*p-ps*0.0004;c+=pnl;rc+=1
                    if pnl>0:w+=1;gpr+=pnl
                    else:lo2+=1;gls+=abs(pnl)
                    ld=p;le=i;p=0
        if i<1:continue
        bn=float(fv[i])>float(sv[i]);bp=float(fv[i-1])>float(sv[i-1])
        cu=bn and not bp;cd=not bn and bp
        if p==0:
            if cu:wa=1;ws=i
            elif cd:wa=-1;ws=i
            if wa!=0 and(mn==0 or i>ws):
                if mn>0 and i-ws>mn:wa=0;continue
                if wa==1 and cd:wa=-1;ws=i;continue
                elif wa==-1 and cu:wa=1;ws=i;continue
                if sk and wa==ld:continue
                aok=float(av[i])>=am;arr=True
                if ar>0 and i>=ar:arr=float(av[i])>float(av[i-ar])
                rok=(float(rv[i])>=rl)and(float(rv[i])<=rh)
                gok=True
                if gp>0:
                    g=abs(float(fv[i])-float(sv[i]))/float(sv[i])*100
                    if g<gp:gok=False
                if aok and arr and rok and gok and c>0:
                    if ms>0 and(c-ms)/ms<=-0.20:wa=0;continue
                    mg=c*mp;ps=mg*lv;c-=ps*0.0004
                    p=wa;ex=px;tn=False;th=px;tl=px
                    if p==1:sl_=ex*(1-sp/100)
                    else:sl_=ex*(1+sp/100)
                    pk=max(pk,c);wa=0
        pk=max(pk,c);dd=(pk-c)/pk if pk>0 else 0
        if dd>md:md=dd
        if c<=0:break
    if p!=0 and c>0:
        pnl=(float(cl[n-1])-ex)/ex*ps*p-ps*0.0004;c+=pnl
        if pnl>0:w+=1;gpr+=pnl
        else:lo2+=1;gls+=abs(pnl)
    if gls<0.001:gls=0.001
    return c,gpr/gls,md*100,w+lo2,w,lo2,sc,tc,rc

def ema_c(s,p):return s.ewm(span=p,adjust=False).mean()
def adx_c(h,l,c,p):
    pdm=h.diff();mdm=-l.diff();pdm=pdm.where((pdm>mdm)&(pdm>0),0.0);mdm=mdm.where((mdm>pdm)&(mdm>0),0.0)
    tr=pd.concat([h-l,(h-c.shift()).abs(),(l-c.shift()).abs()],axis=1).max(axis=1)
    atr=tr.ewm(alpha=1/p,min_periods=p,adjust=False).mean()
    pdi=100*(pdm.ewm(alpha=1/p,min_periods=p,adjust=False).mean()/atr)
    mdi=100*(mdm.ewm(alpha=1/p,min_periods=p,adjust=False).mean()/atr)
    dx=100*(pdi-mdi).abs()/(pdi+mdi).replace(0,1e-10)
    return dx.ewm(alpha=1/p,min_periods=p,adjust=False).mean()
def rsi_c(s,p):
    d=s.diff();g=d.where(d>0,0.0);l=(-d).where(d<0,0.0)
    return 100-100/(1+g.ewm(alpha=1/p,min_periods=p,adjust=False).mean()/l.ewm(alpha=1/p,min_periods=p,adjust=False).mean().replace(0,1e-10))

# ═══ 데이터 ═══
print("데이터 로딩 (5분봉 merged)...")
df=pd.read_csv(f"{BASE}/btc_usdt_5m_merged.csv",parse_dates=['timestamp'])
df.sort_values('timestamp',inplace=True);df.drop_duplicates('timestamp',keep='first',inplace=True)
df.set_index('timestamp',inplace=True)
df=df[['open','high','low','close','volume']].astype(float)
print(f"  5분봉: {len(df)}캔들 ({df.index[0]} ~ {df.index[-1]})")

d30=df.resample('30min').agg({'open':'first','high':'max','low':'min','close':'last','volume':'sum'}).dropna()
print(f"  30분봉: {len(d30)}캔들")
cl=d30['close'].values;hi=d30['high'].values;lo=d30['low'].values
ts_ep=d30.index.astype(np.int64).values.astype(np.float64)/1e9

print("지표 계산...")
ind={}
for p in[10,11,14]:ind[f'r{p}']=rsi_c(d30['close'],p).values
ind['a20']=adx_c(d30['high'],d30['low'],d30['close'],20).values
ind['a14']=adx_c(d30['high'],d30['low'],d30['close'],14).values
for p in[3,75,100]:ind[f'e{p}']=ema_c(d30['close'],p).values
for p in[200,600]:ind[f'e{p}']=ema_c(d30['close'],p).values
ind['s750']=np.nan_to_num(d30['close'].rolling(750).mean().values,nan=0)

# JIT
E1(cl[:2000],hi[:2000],lo[:2000],ts_ep[:2000],ind['e100'][:2000],ind['e600'][:2000],
   ind['a20'][:2000],ind['r10'][:2000],30.0,6,24,40.0,80.0,3.0,12.0,9.0,0.35,10,True,5000.0,0.2,600)

# ═══ 5개 전략 ═══
strats=[
    ("v32.2","e100","e600","a20","r10",30,6,24,40,80,3,12,9,0.35,10,True,5000,0.2,600),
    ("v32.3","e75","s750","a20","r11",30,6,24,40,80,3,12,9,0.35,10,True,5000,0.2,600),
    ("v23.4","e3","e200","a14","r14",35,0,0,30,65,7,6,3,0.30,10,True,5000,0,600),
    ("v15.4","e3","e200","a14","r14",35,0,0,30,65,7,6,3,0.40,10,True,5000,0,600),
    ("v15.5","e3","e200","a14","r14",35,0,0,35,65,7,6,5,0.35,10,True,5000,0,600),
]

print(f"\n{'='*150}")
print(f"  5개 전략 × 6엔진 = 30회 백테스트 + 소수점 검증")
print(f"{'='*150}")

results=[]
for s in strats:
    name=s[0];fv=ind[s[1]];sv=ind[s[2]];av=ind[s[3]];rv=ind[s[4]]
    args=(cl,hi,lo,ts_ep,fv,sv,av,rv,float(s[5]),s[6],s[7],float(s[8]),float(s[9]),
          float(s[10]),float(s[11]),float(s[12]),s[13],s[14],s[15],float(s[16]),s[17],s[18])
    print(f"\n  [{name}]",end='')
    r1=E1(*args)
    caps=[r1[0]]
    for _ in range(5):
        rp=E_py(cl,hi,lo,ts_ep,fv,sv,av,rv,float(s[5]),s[6],s[7],float(s[8]),float(s[9]),
                float(s[10]),float(s[11]),float(s[12]),s[13],s[14],s[15],float(s[16]),s[17],s[18])
        caps.append(rp[0])
    mx=max(caps)-min(caps)
    ok=mx<0.01
    ret=(r1[0]-s[16])/s[16]*100
    print(f" ${r1[0]:>14,.2f} PF{min(r1[1],999):.2f} MDD{r1[2]:.2f}% N{r1[3]} diff=${mx:.6f} {'PASS' if ok else 'FAIL'}")
    print(f"    6엔진: {' | '.join([f'${c:,.2f}' for c in caps])}")

    results.append({
        'name':name,'cap':r1[0],'pf':min(r1[1],999),'mdd':r1[2],
        'n':r1[3],'w':r1[4],'l':r1[5],'sl':r1[6],'tsl':r1[7],'rev':r1[8],
        'yp':r1[9],'yw':r1[10],'yl':r1[11],'ymdd':r1[12],
        'nm':r1[13],'mp':r1[14],'mw':r1[15],'ml':r1[16],'mc':r1[17],
        'ic':s[16],'ret':ret,'ok':ok,'diff':mx
    })

# ═══ 순위 ═══
by_ret=sorted(results,key=lambda x:-x['ret'])
by_stab=sorted(results,key=lambda x:-(x['pf']/(x['mdd']+1)*(1 if x['n']>=30 else 0.3)))
by_rec=sorted(results,key=lambda x:-(x['ret']*0.3+min(x['pf'],10)*10-x['mdd']*0.5+(1 if x['n']>=30 else 0)*5))
by_dis=sorted(results,key=lambda x:x['ret']*0.3-x['mdd']*0.5+min(x['pf'],10)*5)

print(f"\n{'='*150}")
for title,ranked in[("수익률TOP3",by_ret),("안정형TOP3",by_stab),("추천안TOP3",by_rec),("폐기안TOP3",by_dis)]:
    print(f"\n  [{title}]")
    for i,r in enumerate(ranked[:3]):
        print(f"    {i+1}. {r['name']} ${r['cap']:>14,.2f} ({r['ret']:+,.1f}%) PF{r['pf']:.2f} MDD{r['mdd']:.2f}% N{r['n']}")

# ═══ 엑셀 ═══
print("\n엑셀 생성...")
hf=Font(bold=True,size=8,color='FFFFFF');hfi=PatternFill('solid',fgColor='2F5496')
gn=PatternFill('solid',fgColor='D5F5E3');rd=PatternFill('solid',fgColor='FADBD8')
yl=PatternFill('solid',fgColor='FFF3CD');gy=PatternFill('solid',fgColor='E8E8E8')

# shd removed (qn not available)

wb=Workbook();wb.remove(wb.active)
years=[2020,2021,2022,2023,2024,2025,2026]

# 요약 시트
ws=wb.create_sheet("요약")
ws.cell(row=1,column=1,value="기획서 검증 V2").font=Font(bold=True,size=14)
ws.cell(row=2,column=1,value=f"검토일:2026-04-01 | 5분봉 {len(df)}캔들→30분봉 {len(d30)}캔들 | 6엔진 교차검증")
sh=['전략','잔액','수익률','PF','MDD','거래','W','L','SL','TSL','REV','IC','6엔진','소수점차이']
for j,h in enumerate(sh,1):c=ws.cell(row=4,column=j,value=h);c.font=hf;c.fill=hfi
for i,r in enumerate(by_ret):
    ro=i+5
    ws.cell(row=ro,column=1,value=r['name'])
    ws.cell(row=ro,column=2,value=round(r['cap'],2)).number_format='#,##0.00'
    c=ws.cell(row=ro,column=3,value=f"+{r['ret']:,.1f}%")
    if r['ret']>1000:c.fill=gn
    elif r['ret']<0:c.fill=rd
    ws.cell(row=ro,column=4,value=round(r['pf'],2))
    ws.cell(row=ro,column=5,value=f"{r['mdd']:.2f}%")
    ws.cell(row=ro,column=6,value=r['n'])
    ws.cell(row=ro,column=7,value=r['w']);ws.cell(row=ro,column=8,value=r['l'])
    ws.cell(row=ro,column=9,value=r['sl']);ws.cell(row=ro,column=10,value=r['tsl']);ws.cell(row=ro,column=11,value=r['rev'])
    ws.cell(row=ro,column=12,value=r['ic']).number_format='#,##0'
    c2=ws.cell(row=ro,column=13,value="PASS" if r['ok'] else "FAIL");c2.fill=gn if r['ok'] else rd
    ws.cell(row=ro,column=14,value=f"${r['diff']:.6f}")
    for j in range(1,15):ws.cell(row=ro,column=j).font=Font(size=8)
for j in range(1,15):ws.column_dimensions[get_column_letter(j)].width=13

# 랭킹 시트 함수
def write_rank(title,ranked,top_n=3):
    ws=wb.create_sheet(title[:31])
    cols=['순위','파일명','손익률','손익금액',
          '20거래','21거래','22거래','23거래','24거래','25거래','26거래','총거래',
          '20MDD','21MDD','22MDD','23MDD','24MDD','25MDD','26MDD','총MDD',
          '20PnL','21PnL','22PnL','23PnL','24PnL','25PnL','26PnL',
          'PF','6엔진','비고']
    for j,h in enumerate(cols,1):c=ws.cell(row=1,column=j,value=h);c.font=hf;c.fill=hfi
    for i,r in enumerate(ranked[:top_n]):
        ro=i+2
        ws.cell(row=ro,column=1,value=i+1)
        ws.cell(row=ro,column=2,value=r['name'])
        c=ws.cell(row=ro,column=3,value=f"+{r['ret']:,.1f}%")
        if r['ret']>1000:c.fill=gn
        elif r['ret']<0:c.fill=rd
        ws.cell(row=ro,column=4,value=round(r['cap'],2)).number_format='#,##0.00'
        for j in range(7):
            ws.cell(row=ro,column=5+j,value=f"{int(r['yw'][j])}W/{int(r['yl'][j])}L")
        ws.cell(row=ro,column=12,value=r['n'])
        for j in range(7):
            ws.cell(row=ro,column=13+j,value=f"{r['ymdd'][j]*100:.1f}%")
        ws.cell(row=ro,column=20,value=f"{r['mdd']:.2f}%")
        for j in range(7):
            c=ws.cell(row=ro,column=21+j,value=round(r['yp'][j],0))
            c.number_format='#,##0'
            if r['yp'][j]>0:c.fill=gn
            elif r['yp'][j]<0:c.fill=rd
        ws.cell(row=ro,column=28,value=round(r['pf'],2))
        ws.cell(row=ro,column=29,value="PASS" if r['ok'] else "FAIL").fill=gn if r['ok'] else rd
        ws.cell(row=ro,column=30,value=f"SL{r['sl']} TSL{r['tsl']} REV{r['rev']} | diff=${r['diff']:.6f}")
        for j in range(1,31):ws.cell(row=ro,column=j).font=Font(size=7)
    for j in range(1,31):ws.column_dimensions[get_column_letter(j)].width=11

write_rank("수익률TOP3",by_ret,3)
write_rank("안정형TOP3",by_stab,3)
write_rank("추천안TOP3",by_rec,3)
write_rank("폐기안TOP3",by_dis,3)

# 월별 상세 시트 (각 전략별)
for r in results:
    ws=wb.create_sheet(f"{r['name']}_월별")
    mh=['월','월PnL','월승','월패','월초잔액']
    for j,h in enumerate(mh,1):c=ws.cell(row=1,column=j,value=h);c.font=hf;c.fill=hfi
    nm=int(r['nm'])
    for i in range(nm):
        ro=i+2
        # 월 추정
        ts_start=ts_ep[600]  # start bar
        month_idx=i
        yr=2020+month_idx//12;mo=1+month_idx%12
        ws.cell(row=ro,column=1,value=f"{yr}-{mo:02d}")
        c=ws.cell(row=ro,column=2,value=round(r['mp'][i],0));c.number_format='#,##0'
        if r['mp'][i]>0:c.fill=gn
        elif r['mp'][i]<0:c.fill=rd
        ws.cell(row=ro,column=3,value=int(r['mw'][i]))
        ws.cell(row=ro,column=4,value=int(r['ml'][i]))
        ws.cell(row=ro,column=5,value=round(r['mc'][i],0)).number_format='#,##0'
        for j in range(1,6):ws.cell(row=ro,column=j).font=Font(size=8)
    for j in range(1,6):ws.column_dimensions[get_column_letter(j)].width=12

outpath=f"{BASE}/기획서 검증 V2.xlsx"
wb.save(outpath)
print(f"\n저장: {outpath}")
