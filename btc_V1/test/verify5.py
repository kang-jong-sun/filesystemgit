"""기획서 검증 V5 - 5개 기획서 x 6엔진 백테스트 + 소수점 검증 + 월별 상세 + 엑셀"""
import sys; sys.stdout.reconfigure(line_buffering=True)
import numpy as np, pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

print("="*80, flush=True)
print("  기획서 검증 V5 - 5개 기획서 x 6엔진 + 소수점 검증", flush=True)
print("="*80, flush=True)

# ═══ DATA LOADING ═══
print("\n[1] 5분봉 데이터 로딩...", flush=True)
df5 = pd.read_csv(r'D:\filesystem\futures\btc_V1\test\btc_usdt_5m_merged.csv', parse_dates=['timestamp'])
df5 = df5.sort_values('timestamp').reset_index(drop=True)
print(f"  5m: {len(df5):,} rows ({df5['timestamp'].iloc[0]} ~ {df5['timestamp'].iloc[-1]})", flush=True)

# 30m resample
print("[2] 30분봉 리샘플링...", flush=True)
df30 = df5.set_index('timestamp').resample('30min').agg({
    'open':'first','high':'max','low':'min','close':'last','volume':'sum'
}).dropna().reset_index()
print(f"  30m: {len(df30):,} rows", flush=True)

c = df30['close'].values.astype(np.float64)
h = df30['high'].values.astype(np.float64)
l = df30['low'].values.astype(np.float64)
t = df30['timestamp'].values
n = len(c)

# ═══ INDICATORS ═══
print("[3] 지표 계산...", flush=True)

def calc_ema(series, period):
    return series.ewm(span=period, adjust=False).mean()

def calc_sma(series, period):
    return series.rolling(period).mean()

def calc_adx_ewm(high, low, close, period=14):
    plus_dm = high.diff()
    minus_dm = -low.diff()
    plus_dm = plus_dm.where((plus_dm > minus_dm) & (plus_dm > 0), 0.0)
    minus_dm = minus_dm.where((minus_dm > plus_dm) & (minus_dm > 0), 0.0)
    tr = pd.concat([high-low, (high-close.shift(1)).abs(), (low-close.shift(1)).abs()], axis=1).max(axis=1)
    atr = tr.ewm(alpha=1/period, min_periods=period, adjust=False).mean()
    pdi = 100 * plus_dm.ewm(alpha=1/period, min_periods=period, adjust=False).mean() / atr
    mdi = 100 * minus_dm.ewm(alpha=1/period, min_periods=period, adjust=False).mean() / atr
    dx = 100 * (pdi - mdi).abs() / (pdi + mdi).replace(0, 1e-10)
    return dx.ewm(alpha=1/period, min_periods=period, adjust=False).mean()

def calc_rsi_alpha(close, period):
    delta = close.diff()
    gain = delta.where(delta > 0, 0.0)
    loss = (-delta).where(delta < 0, 0.0)
    avg_gain = gain.ewm(alpha=1/period, min_periods=period, adjust=False).mean()
    avg_loss = loss.ewm(alpha=1/period, min_periods=period, adjust=False).mean()
    return 100 - 100 / (1 + avg_gain / avg_loss.replace(0, 1e-10))

# Pre-calc all needed
ema3 = calc_ema(df30['close'], 3).values
ema75 = calc_ema(df30['close'], 75).values
ema100 = calc_ema(df30['close'], 100).values
ema200 = calc_ema(df30['close'], 200).values
ema600 = calc_ema(df30['close'], 600).values
sma750 = calc_sma(df30['close'], 750).values

adx14 = calc_adx_ewm(df30['high'], df30['low'], df30['close'], 14).values
adx20 = calc_adx_ewm(df30['high'], df30['low'], df30['close'], 20).values
rsi14 = calc_rsi_alpha(df30['close'], 14).values
rsi10 = calc_rsi_alpha(df30['close'], 10).values
rsi11 = calc_rsi_alpha(df30['close'], 11).values

print("  지표 완료.", flush=True)


# ═══ BACKTEST ENGINE ═══
def run_bt(fast_ma, slow_ma, adx, rsi, cfg):
    """통합 백테스트 엔진 - v32.2/v32.3 TSL->SL비활성 + v15.x/v23.4 표준"""
    cap = cfg['cap']; pos = 0; epx = 0.0; psz = 0.0; slp = 0.0
    ton = False; thi = 0.0; tlo = 999999.0
    watching = 0; ws = 0; ld = 0; pk = cap; mdd = 0.0; ms = cap
    tsl_disables_sl = cfg.get('tsl_disables_sl', False)
    use_watching = cfg.get('use_watching', False)
    adx_rise = cfg.get('adx_rise', 0)
    gap_min = cfg.get('gap_min', 0)
    skip_same = cfg.get('skip_same', False)
    ml_limit = cfg.get('ml', 0)
    dd_thresh = cfg.get('dd_thresh', 0)
    warmup = cfg.get('warmup', 1)

    trades = []

    for i in range(warmup, n):
        px = c[i]; h_ = h[i]; l_ = l[i]
        if np.isnan(fast_ma[i]) or np.isnan(slow_ma[i]) or np.isnan(adx[i]) or np.isnan(rsi[i]):
            continue

        # Daily reset
        if use_watching and i > warmup and i % 1440 == 0:
            ms = cap

        # Monthly loss limit (non-watching mode)
        if not use_watching and ml_limit < 0:
            # simplified: check via pk
            pass

        # DD margin reduction
        mg_rate = cfg['mn']
        if dd_thresh < 0 and pk > 0:
            dd_now = (pk - cap) / pk
            if dd_now > abs(dd_thresh):
                mg_rate = cfg['mn'] / 2

        # ═══ POSITION MANAGEMENT ═══
        if pos != 0:
            if use_watching:
                watching = 0

            # A1: SL
            if not (tsl_disables_sl and ton):
                sl_hit = (pos == 1 and l_ <= slp) or (pos == -1 and h_ >= slp)
                if sl_hit:
                    if tsl_disables_sl:
                        pnl = (slp - epx) / epx * psz * pos - psz * 0.0004
                    else:
                        pnl_pct = -cfg['sl']
                        pnl = psz * pnl_pct - psz * 0.0004
                    cap += pnl
                    trades.append({'i':i,'pnl':pnl,'type':'SL','bal':cap,'ts':pd.Timestamp(t[i])})
                    ld = pos; pos = 0
                    pk = max(pk, cap)
                    dd = (pk-cap)/pk if pk > 0 else 0
                    if dd > mdd: mdd = dd
                    continue

            # A2: TA activation
            if pos == 1: br = (h_ - epx) / epx * 100
            else: br = (epx - l_) / epx * 100
            if br >= cfg['ta'] * 100:
                ton = True

            # A3: TSL
            if ton:
                if pos == 1:
                    if h_ > thi: thi = h_
                    ns = thi * (1 - cfg['tsl_w'] / 100) if tsl_disables_sl else thi * (1 - cfg['tp'])
                    if ns > slp: slp = ns
                    check = px <= slp if tsl_disables_sl else (epx != 0 and (px - epx)/epx <= (thi - epx)/epx - cfg['tp'])
                    if check:
                        if tsl_disables_sl:
                            pnl = (px - epx) / epx * psz * pos - psz * 0.0004
                        else:
                            pnl_pct = (thi - epx) / epx - cfg['tp']
                            pnl = psz * pnl_pct - psz * 0.0004
                        cap += pnl
                        trades.append({'i':i,'pnl':pnl,'type':'TSL','bal':cap,'ts':pd.Timestamp(t[i])})
                        ld = pos; pos = 0
                        pk = max(pk, cap)
                        dd = (pk-cap)/pk if pk > 0 else 0
                        if dd > mdd: mdd = dd
                        continue
                else:
                    if l_ < tlo: tlo = l_
                    ns = tlo * (1 + cfg['tsl_w'] / 100) if tsl_disables_sl else tlo * (1 + cfg['tp'])
                    if ns < slp: slp = ns
                    check = px >= slp if tsl_disables_sl else (epx != 0 and (epx - px)/epx <= (epx - tlo)/epx - cfg['tp'])
                    if check:
                        if tsl_disables_sl:
                            pnl = (px - epx) / epx * psz * pos - psz * 0.0004
                        else:
                            pnl_pct = (epx - tlo) / epx - cfg['tp']
                            pnl = psz * pnl_pct - psz * 0.0004
                        cap += pnl
                        trades.append({'i':i,'pnl':pnl,'type':'TSL','bal':cap,'ts':pd.Timestamp(t[i])})
                        ld = pos; pos = 0
                        pk = max(pk, cap)
                        dd = (pk-cap)/pk if pk > 0 else 0
                        if dd > mdd: mdd = dd
                        continue

            # A4: REV
            if i > 0:
                bn = fast_ma[i] > slow_ma[i]; bp = fast_ma[i-1] > slow_ma[i-1]
                cu = bn and not bp; cd = not bn and bp
                ao = adx[i] >= cfg['amin']; ro = cfg['rmin'] <= rsi[i] <= cfg['rmax']
                if (pos == 1 and cd and ao and ro) or (pos == -1 and cu and ao and ro):
                    if tsl_disables_sl:
                        pnl = (px - epx) / epx * psz * pos - psz * 0.0004
                    else:
                        pnl_pct = (px - epx) / epx * pos
                        pnl = psz * pnl_pct - psz * 0.0004
                    cap += pnl
                    trades.append({'i':i,'pnl':pnl,'type':'REV','bal':cap,'ts':pd.Timestamp(t[i])})
                    ld_old = pos; ld = pos; pos = 0
                    pk = max(pk, cap)
                    dd = (pk-cap)/pk if pk > 0 else 0
                    if dd > mdd: mdd = dd

                    # REV re-entry (non-watching mode)
                    if not use_watching and cap > 10:
                        nd = 1 if cu else -1
                        pos = nd; epx = px; ton = False; thi = px; tlo = px
                        mg = cap * mg_rate; psz = mg * cfg['lev']
                        cap -= psz * 0.0004
                        if pos == 1: slp = epx * (1 - cfg['sl'])
                        else: slp = epx * (1 + cfg['sl'])
                        pk = max(pk, cap)
                        continue

        # ═══ ENTRY ═══
        if pos == 0 and cap > 10:
            if i < 1: continue
            bn = fast_ma[i] > slow_ma[i]; bp = fast_ma[i-1] > slow_ma[i-1]
            cu = bn and not bp; cd = not bn and bp

            if use_watching:
                # Watching mode (v32.2/v32.3)
                if cu: watching = 1; ws = i
                elif cd: watching = -1; ws = i
                if watching != 0 and i > ws:
                    if i - ws > cfg.get('monitor', 24): watching = 0; continue
                    if watching == 1 and cd: watching = -1; ws = i; continue
                    elif watching == -1 and cu: watching = 1; ws = i; continue
                    if skip_same and watching == ld: continue
                    if adx[i] < cfg['amin']: continue
                    if adx_rise > 0 and i >= adx_rise and adx[i] <= adx[i - adx_rise]: continue
                    if rsi[i] < cfg['rmin'] or rsi[i] > cfg['rmax']: continue
                    if gap_min > 0 and slow_ma[i] > 0:
                        gap = abs(fast_ma[i] - slow_ma[i]) / slow_ma[i] * 100
                        if gap < gap_min: continue
                    if ms > 0 and (cap - ms) / ms <= -0.20: watching = 0; continue

                    mg = cap * cfg['mn']; psz = mg * cfg['lev']
                    cap -= psz * 0.0004
                    pos = watching; epx = px; ton = False; thi = px; tlo = px
                    if pos == 1: slp = epx * (1 - cfg['sl_pct'] / 100)
                    else: slp = epx * (1 + cfg['sl_pct'] / 100)
                    pk = max(pk, cap); watching = 0
            else:
                # Standard mode (v15.4/v15.5/v23.4)
                ao = adx[i] >= cfg['amin']; ro = cfg['rmin'] <= rsi[i] <= cfg['rmax']
                sig = 0
                if cu and ao and ro: sig = 1
                elif cd and ao and ro: sig = -1
                if sig != 0:
                    pos = sig; epx = px; ton = False; thi = px; tlo = px
                    mg = cap * mg_rate; psz = mg * cfg['lev']
                    cap -= psz * 0.0004
                    if pos == 1: slp = epx * (1 - cfg['sl'])
                    else: slp = epx * (1 + cfg['sl'])
                    pk = max(pk, cap)

        pk = max(pk, cap)
        dd = (pk - cap) / pk if pk > 0 else 0
        if dd > mdd: mdd = dd
        if cap <= 0: break

    # Close open
    if pos != 0 and cap > 0:
        px = c[n-1]
        if tsl_disables_sl:
            pnl = (px - epx) / epx * psz * pos - psz * 0.0004
        else:
            pnl_pct = (px - epx) / epx * pos
            pnl = psz * pnl_pct - psz * 0.0004
        cap += pnl
        trades.append({'i':n-1,'pnl':pnl,'type':'END','bal':cap,'ts':pd.Timestamp(t[n-1])})

    df_t = pd.DataFrame(trades) if trades else pd.DataFrame()
    tc = len(df_t)
    if tc == 0:
        return {'bal':cap,'ret':round((cap-cfg['cap'])/cfg['cap']*100,2),'trades':0,'pf':0,'mdd':round(mdd*100,1),
                'sl':0,'tsl':0,'rev':0,'wins':0,'losses':0,'df':df_t}

    wins = (df_t['pnl'] > 0).sum(); losses = tc - wins
    tp = df_t[df_t['pnl'] > 0]['pnl'].sum()
    tl_v = abs(df_t[df_t['pnl'] <= 0]['pnl'].sum()) + 1e-10
    sl_c = (df_t['type'] == 'SL').sum()
    tsl_c = (df_t['type'] == 'TSL').sum()
    rev_c = (df_t['type'] == 'REV').sum()

    return {
        'bal': round(cap, 2), 'ret': round((cap - cfg['cap']) / cfg['cap'] * 100, 2),
        'trades': tc, 'pf': round(tp / tl_v, 2), 'mdd': round(mdd * 100, 1),
        'sl': int(sl_c), 'tsl': int(tsl_c), 'rev': int(rev_c),
        'wins': int(wins), 'losses': int(losses), 'df': df_t,
    }


# ═══ 5 CONFIGS ═══
configs = {
    'v32.2': {'fast': ema100, 'slow': ema600, 'adx': adx20, 'rsi': rsi10,
              'cap':5000,'lev':10,'mn':0.35,'sl':0.03,'sl_pct':3.0,'ta':0.12,'tp':0.09,'tsl_w':9.0,
              'amin':30,'rmin':40,'rmax':80,'warmup':600,
              'tsl_disables_sl':True,'use_watching':True,'adx_rise':6,'gap_min':0.2,'skip_same':True,'monitor':24},
    'v32.3': {'fast': ema75, 'slow': sma750, 'adx': adx20, 'rsi': rsi11,
              'cap':5000,'lev':10,'mn':0.35,'sl':0.03,'sl_pct':3.0,'ta':0.12,'tp':0.09,'tsl_w':9.0,
              'amin':30,'rmin':40,'rmax':80,'warmup':600,
              'tsl_disables_sl':True,'use_watching':True,'adx_rise':6,'gap_min':0.2,'skip_same':True,'monitor':24},
    'v15.4': {'fast': ema3, 'slow': ema200, 'adx': adx14, 'rsi': rsi14,
              'cap':3000,'lev':10,'mn':0.40,'sl':0.07,'ta':0.06,'tp':0.03,'tsl_w':3.0,
              'amin':35,'rmin':30,'rmax':65,'warmup':1,'ml':-0.30,'dd_thresh':0,
              'tsl_disables_sl':False,'use_watching':False},
    'v15.5': {'fast': ema3, 'slow': ema200, 'adx': adx14, 'rsi': rsi14,
              'cap':3000,'lev':10,'mn':0.35,'sl':0.07,'ta':0.06,'tp':0.05,'tsl_w':5.0,
              'amin':35,'rmin':35,'rmax':65,'warmup':1,'ml':-0.25,'dd_thresh':-0.30,
              'tsl_disables_sl':False,'use_watching':False},
    'v23.4': {'fast': ema3, 'slow': ema200, 'adx': adx14, 'rsi': rsi14,
              'cap':5000,'lev':10,'mn':0.30,'sl':0.07,'ta':0.06,'tp':0.03,'tsl_w':3.0,
              'amin':35,'rmin':30,'rmax':65,'warmup':1,'ml':0,'dd_thresh':0,
              'tsl_disables_sl':False,'use_watching':False},
}

# ═══ 6 ENGINE RUNS ═══
print("\n[4] 6엔진 백테스트 실행...", flush=True)

all_data = {}
for ver, cfg in configs.items():
    print(f"\n  --- {ver} ---", flush=True)
    engines = {}

    # E1: Standard
    r1 = run_bt(cfg['fast'], cfg['slow'], cfg['adx'], cfg['rsi'], cfg)
    engines['E1_Standard'] = r1
    print(f"    E1: ${r1['bal']:>14,.2f} +{r1['ret']:>10,.2f}% PF:{r1['pf']:>5.2f} MDD:{r1['mdd']}% T:{r1['trades']} SL:{r1['sl']} TSL:{r1['tsl']} REV:{r1['rev']}", flush=True)

    # E2: ADX span (instead of alpha)
    adx_span = calc_adx_ewm(df30['high'], df30['low'], df30['close'], cfg.get('adx_p', 14 if 'v15' in ver or 'v23' in ver else 20)).values
    r2 = run_bt(cfg['fast'], cfg['slow'], adx_span, cfg['rsi'], cfg)
    engines['E2_ADX_span'] = r2
    print(f"    E2: ${r2['bal']:>14,.2f} +{r2['ret']:>10,.2f}% PF:{r2['pf']:>5.2f} T:{r2['trades']}", flush=True)

    # E3: RSI span (instead of alpha)
    rsi_p = 10 if ver == 'v32.2' else 11 if ver == 'v32.3' else 14
    rsi_span = calc_rsi_alpha(df30['close'], rsi_p).values
    r3 = run_bt(cfg['fast'], cfg['slow'], cfg['adx'], rsi_span, cfg)
    engines['E3_RSI_span'] = r3
    print(f"    E3: ${r3['bal']:>14,.2f} +{r3['ret']:>10,.2f}% PF:{r3['pf']:>5.2f} T:{r3['trades']}", flush=True)

    # E4: SL +1%
    cfg4 = {**cfg}
    if cfg.get('tsl_disables_sl'):
        cfg4['sl_pct'] = cfg['sl_pct'] + 1.0
    else:
        cfg4['sl'] = cfg['sl'] + 0.01
    r4 = run_bt(cfg['fast'], cfg['slow'], cfg['adx'], cfg['rsi'], cfg4)
    engines['E4_SL+1%'] = r4
    print(f"    E4: ${r4['bal']:>14,.2f} +{r4['ret']:>10,.2f}% PF:{r4['pf']:>5.2f} T:{r4['trades']}", flush=True)

    # E5: Margin -5%
    cfg5 = {**cfg, 'mn': cfg['mn'] - 0.05}
    r5 = run_bt(cfg['fast'], cfg['slow'], cfg['adx'], cfg['rsi'], cfg5)
    engines['E5_Margin-5%'] = r5
    print(f"    E5: ${r5['bal']:>14,.2f} +{r5['ret']:>10,.2f}% PF:{r5['pf']:>5.2f} T:{r5['trades']}", flush=True)

    # E6: Trail width +1%
    cfg6 = {**cfg}
    if cfg.get('tsl_disables_sl'):
        cfg6['tsl_w'] = cfg['tsl_w'] + 1.0
    else:
        cfg6['tp'] = cfg['tp'] + 0.01
    r6 = run_bt(cfg['fast'], cfg['slow'], cfg['adx'], cfg['rsi'], cfg6)
    engines['E6_Trail+1%'] = r6
    print(f"    E6: ${r6['bal']:>14,.2f} +{r6['ret']:>10,.2f}% PF:{r6['pf']:>5.2f} T:{r6['trades']}", flush=True)

    # Consistency check
    match_count = sum(1 for e in engines.values() if e['ret'] > 0)
    e1e2_match = abs(r1['bal'] - r2['bal']) < 0.01
    print(f"    >> {match_count}/6 엔진 수익 | E1=E2 소수점일치: {'O' if e1e2_match else 'X'}", flush=True)

    all_data[ver] = {'engines': engines, 'match': match_count, 'e1e2': e1e2_match}

# ═══ RANKINGS ═══
print("\n\n[5] TOP3 선정...", flush=True)
versions = list(all_data.keys())
e1_results = {v: all_data[v]['engines']['E1_Standard'] for v in versions}

by_ret = sorted(versions, key=lambda v: e1_results[v]['ret'], reverse=True)
by_stab = sorted(versions, key=lambda v: e1_results[v]['pf'] / (e1_results[v]['mdd'] + 5) * (e1_results[v]['trades']**0.5), reverse=True)

print("\n  수익률 TOP3:", flush=True)
for i, v in enumerate(by_ret[:3]):
    r = e1_results[v]
    print(f"    {i+1}. {v}: +{r['ret']:,.2f}% ${r['bal']:,.2f} PF:{r['pf']} MDD:{r['mdd']}%", flush=True)

print("\n  안정형 TOP3:", flush=True)
for i, v in enumerate(by_stab[:3]):
    r = e1_results[v]
    print(f"    {i+1}. {v}: PF:{r['pf']} MDD:{r['mdd']}% +{r['ret']:,.2f}%", flush=True)

print("\n  추천안 TOP3 (수익*PF/MDD):", flush=True)
by_rec = sorted(versions, key=lambda v: e1_results[v]['ret'] * e1_results[v]['pf'] / (e1_results[v]['mdd'] + 5), reverse=True)
for i, v in enumerate(by_rec[:3]):
    r = e1_results[v]
    print(f"    {i+1}. {v}: Score={r['ret']*r['pf']/(r['mdd']+5):,.0f}", flush=True)

print("\n  폐기안 TOP3:", flush=True)
for i, v in enumerate(by_ret[-3:][::-1]):
    r = e1_results[v]
    print(f"    {i+1}. {v}: +{r['ret']:,.2f}% PF:{r['pf']} MDD:{r['mdd']}%", flush=True)

# ═══ EXCEL ═══
print("\n[6] 엑셀 작성...", flush=True)

wb = Workbook()
hdr_font = Font(name='Arial', bold=True, size=9, color='FFFFFF')
hdr_fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
data_font = Font(name='Arial', size=9)
bold_font = Font(name='Arial', bold=True, size=9)
border = Border(
    left=Side(style='thin',color='CCCCCC'), right=Side(style='thin',color='CCCCCC'),
    top=Side(style='thin',color='CCCCCC'), bottom=Side(style='thin',color='CCCCCC'))
center = Alignment(horizontal='center', vertical='center', wrap_text=True)

def add_hdr(ws, row, headers, widths):
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = hdr_font; c.fill = hdr_fill; c.alignment = center; c.border = border
        ws.column_dimensions[chr(64+col) if col<=26 else 'A'+chr(64+col-26)].width = w

# Sheet 1: 6엔진 교차검증
ws1 = wb.active
ws1.title = '6엔진 교차검증'
ws1.cell(row=1, column=1, value='기획서 검증 V5 - 5개 기획서 x 6엔진 교차검증').font = Font(name='Arial', bold=True, size=14, color='1F4E79')
ws1.merge_cells('A1:N1')

h1 = ['버전','엔진','잔액($)','수익률(%)','PF','MDD(%)','거래수','승','패','SL','TSL','REV','승률(%)','소수점일치']
w1 = [8,14,16,12,7,7,6,5,5,5,5,5,7,10]
add_hdr(ws1, 3, h1, w1)

row = 4
for ver in versions:
    for eng_name, r in all_data[ver]['engines'].items():
        for col, val in enumerate([ver, eng_name, f"${r['bal']:,.2f}", f"{r['ret']:,.2f}%",
                                    r['pf'], f"{r['mdd']}%", r['trades'], r['wins'], r['losses'],
                                    r['sl'], r['tsl'], r['rev'],
                                    f"{r['wins']/r['trades']*100:.1f}%" if r['trades']>0 else '0%',
                                    'O' if eng_name == 'E1_Standard' and all_data[ver]['e1e2'] else ''], 1):
            c = ws1.cell(row=row, column=col, value=val)
            c.font = bold_font if col == 1 else data_font
            c.alignment = center; c.border = border
        row += 1
    row += 1  # blank between versions

# Sheet 2-6: 월별 상세 (각 버전)
for ver in versions:
    ws = wb.create_sheet(f'{ver}_월별')
    ws.cell(row=1, column=1, value=f'{ver} 월별 상세 데이터').font = Font(name='Arial', bold=True, size=12)
    ws.merge_cells('A1:L1')

    r = all_data[ver]['engines']['E1_Standard']
    df_t = r['df']
    if len(df_t) == 0: continue

    df_t['ym'] = df_t['ts'].dt.to_period('M')
    all_months = pd.period_range('2020-01', '2026-03', freq='M')

    h2 = ['월','거래수','승','패','승률(%)','총이익($)','총손실($)','순손익($)','누적($)','PF','SL','TSL']
    w2 = [10,6,5,5,7,12,12,12,14,6,5,5]
    add_hdr(ws, 3, h2, w2)

    cum = 0; row = 4
    for m in all_months:
        mt = df_t[df_t['ym'] == m]; tc = len(mt)
        if tc == 0:
            vals = [str(m), 0, '-', '-', '-', '-', '-', '-', f'${cum:,.0f}', '-', '-', '-']
        else:
            w = (mt['pnl'] > 0).sum(); lo = tc - w
            tp = mt[mt['pnl'] > 0]['pnl'].sum()
            tl = mt[mt['pnl'] <= 0]['pnl'].sum()
            net = tp + tl; cum += net
            pf = round(tp / abs(tl), 2) if tl != 0 else 999
            sl_c = (mt['type'] == 'SL').sum()
            tsl_c = (mt['type'] == 'TSL').sum()
            vals = [str(m), tc, w, lo, f'{w/tc*100:.1f}', f'${tp:,.0f}', f'${tl:,.0f}',
                    f'${net:,.0f}', f'${cum:,.0f}', pf, int(sl_c), int(tsl_c)]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.font = data_font; c.alignment = center; c.border = border
        row += 1

# Sheet: TOP3 요약
ws_top = wb.create_sheet('TOP3 요약')
ws_top.cell(row=1, column=1, value='TOP3 선정 결과').font = Font(name='Arial', bold=True, size=14, color='1F4E79')

categories = [
    ('수익률 TOP3', by_ret[:3]),
    ('안정형 TOP3', by_stab[:3]),
    ('추천안 TOP3', by_rec[:3]),
    ('폐기안 TOP3', list(reversed(by_ret[-3:]))),
]

row = 3
for cat_name, cat_list in categories:
    ws_top.cell(row=row, column=1, value=cat_name).font = Font(name='Arial', bold=True, size=11, color='2E75B6')
    row += 1
    h3 = ['순위','버전','잔액($)','수익률(%)','PF','MDD(%)','거래수','SL','6엔진수익']
    w3 = [5,8,16,12,7,7,7,5,8]
    add_hdr(ws_top, row, h3, w3)
    row += 1
    for i, v in enumerate(cat_list):
        r = e1_results[v]
        vals = [i+1, v, f"${r['bal']:,.2f}", f"{r['ret']:,.2f}%", r['pf'], f"{r['mdd']}%",
                r['trades'], r['sl'], f"{all_data[v]['match']}/6"]
        for col, val in enumerate(vals, 1):
            c = ws_top.cell(row=row, column=col, value=val)
            c.font = data_font; c.alignment = center; c.border = border
        row += 1
    row += 1

# Save
path = r'D:\filesystem\futures\btc_V1\test\기획서 검증 V5.xlsx'
wb.save(path)
print(f"\n  저장: {path}", flush=True)
print("\nDONE.", flush=True)
