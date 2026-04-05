"""
기획서 검증 V1: 5개 기획서 × 6엔진 정밀 교차검증
- 소수점 정밀 검증
- 월별 상세 데이터
- 수익률/안정형/추천/폐기 TOP3
- Excel 출력
"""
import sys,time,os
import pandas as pd,numpy as np
import warnings;warnings.filterwarnings('ignore')
sys.stdout.reconfigure(line_buffering=True)
from openpyxl import Workbook
from openpyxl.styles import Font,PatternFill,Alignment,Border,Side
from openpyxl.utils import get_column_letter

DATA=r"D:\filesystem\futures\btc_V1\test\btc_usdt_5m_merged.csv"
OUT=r"D:\filesystem\futures\btc_V1\test\기획서 검증 V1.xlsx"

# ============ INDICATORS ============
def _ema(d,p):return pd.Series(d).ewm(span=p,adjust=False).mean().values
def _sma(d,p):return pd.Series(d).rolling(p,min_periods=p).mean().values

def adx_wilder(hi,lo,cl,p):
    n=len(cl);tr=np.zeros(n);pdm=np.zeros(n);mdm=np.zeros(n)
    for i in range(1,n):
        tr[i]=max(hi[i]-lo[i],abs(hi[i]-cl[i-1]),abs(lo[i]-cl[i-1]))
        up=hi[i]-hi[i-1];dn=lo[i-1]-lo[i]
        pdm[i]=up if up>dn and up>0 else 0
        mdm[i]=dn if dn>up and dn>0 else 0
    atr=np.zeros(n);pds=np.zeros(n);mds=np.zeros(n)
    if n>p:
        atr[p]=np.mean(tr[1:p+1]);pds[p]=np.mean(pdm[1:p+1]);mds[p]=np.mean(mdm[1:p+1])
        for i in range(p+1,n):
            atr[i]=atr[i-1]*(1-1.0/p)+tr[i]/p
            pds[i]=pds[i-1]*(1-1.0/p)+pdm[i]/p
            mds[i]=mds[i-1]*(1-1.0/p)+mdm[i]/p
    dx=np.zeros(n)
    for i in range(p,n):
        if atr[i]>0:
            pdi=100*pds[i]/atr[i];mdi=100*mds[i]/atr[i];s=pdi+mdi
            dx[i]=100*abs(pdi-mdi)/s if s>0 else 0
    r=np.zeros(n);st=2*p
    if n>st:
        r[st]=np.mean(dx[p:st+1])
        for i in range(st+1,n):r[i]=r[i-1]*(1-1.0/p)+dx[i]/p
    return r

def adx_ewm(hi,lo,cl,p):
    tr1=hi-lo;tr2=np.abs(hi-np.roll(cl,1));tr3=np.abs(lo-np.roll(cl,1))
    tr1[0]=tr2[0]=tr3[0]=0;tr=np.maximum(np.maximum(tr1,tr2),tr3)
    up=hi-np.roll(hi,1);dn=np.roll(lo,1)-lo;up[0]=dn[0]=0
    pdm=np.where((up>dn)&(up>0),up,0.0);mdm=np.where((dn>up)&(dn>0),dn,0.0)
    a=1.0/p
    av=pd.Series(tr).ewm(alpha=a,min_periods=p,adjust=False).mean().values
    ps=pd.Series(pdm).ewm(alpha=a,min_periods=p,adjust=False).mean().values
    ms=pd.Series(mdm).ewm(alpha=a,min_periods=p,adjust=False).mean().values
    with np.errstate(divide='ignore',invalid='ignore'):
        pdi=np.where(av>0,100*ps/av,0);mdi=np.where(av>0,100*ms/av,0)
        ds=pdi+mdi;dx=np.where(ds>0,100*np.abs(pdi-mdi)/ds,0)
    return pd.Series(dx).ewm(alpha=a,min_periods=p,adjust=False).mean().values

def calc_rsi(cl,p):
    d2=np.diff(cl,prepend=cl[0])
    g=np.where(d2>0,d2,0.0);l=np.where(d2<0,-d2,0.0)
    ga=pd.Series(g).ewm(alpha=1.0/p,min_periods=p,adjust=False).mean().values
    la=pd.Series(l).ewm(alpha=1.0/p,min_periods=p,adjust=False).mean().values
    with np.errstate(divide='ignore',invalid='ignore'):
        rs=np.where(la>0,ga/la,100.0)
    return 100-100/(1+rs)


# ============ BACKTEST ENGINE (supports v32.x TSL/SL priority) ============
def backtest(cl,hi,lo,yr,mk,ts_month,
             ma_fast,ma_slow,adx_arr,rsi_arr,
             cfg,engine_cfg):
    """
    cfg: strategy config dict
    engine_cfg: engine variant config dict
    Returns: detailed results with monthly breakdown
    """
    n=len(cl)
    init=cfg['init'];lev=cfg['l'];fee=engine_cfg['fee']
    sl_pct=cfg['sl'];ta_pct=cfg['ta'];tp_pct=cfg['tp']
    adx_min=cfg['at'];rsi_min=cfg['rm'];rsi_max=cfg['rx']
    delay=cfg['d'];mpct=cfg['m']
    ml_limit=cfg.get('ml',-0.20)
    dd_threshold=cfg.get('dd',None)
    dd_margin_reduce=cfg.get('dd_m',None)
    tsl_disables_sl=cfg.get('tsl_disables_sl',False)
    monitor_window=cfg.get('monitor',0)
    adx_rise_bars=cfg.get('adx_rise',0)
    ema_gap_min=cfg.get('ema_gap',0)
    skip_same=engine_cfg['skip']
    strict=engine_cfg['strict']
    wu=cfg.get('wu',300)
    inv=1.0/lev

    bal=init;peak=init;mdd=0.0
    trades=[];in_pos=False;p_dir=0;p_entry=0.0;p_size=0.0;p_margin=0.0
    p_sl=0.0;p_high=0.0;p_low=0.0;tsl_active=False;tsl_price=0.0
    cur_m=-1;m_start=bal;m_locked=False
    pend=0;pcnt=0;cross_bar=-999
    peak_dd=init  # For DD protection

    valid=~(np.isnan(ma_fast)|np.isnan(ma_slow))
    above=np.where(valid,ma_fast>ma_slow,False)

    for i in range(wu,n):
        c=cl[i];h=hi[i];l=lo[i]
        # Month reset
        m=mk[i]
        if m!=cur_m:cur_m=m;m_start=bal;m_locked=False
        if not m_locked and m_start>0:
            if (bal-m_start)/m_start<=ml_limit:m_locked=True

        # DD protection (v15.5)
        actual_margin=mpct
        if dd_threshold and dd_margin_reduce:
            if peak_dd>0 and (peak_dd-bal)/peak_dd>=abs(dd_threshold):
                actual_margin=dd_margin_reduce
            if bal>peak_dd:peak_dd=bal

        if in_pos:
            # --- POSITION MANAGEMENT ---
            # FL check
            if p_dir==1 and l<=p_entry*(1-inv):
                loss=p_margin;bal-=loss
                trades.append({'t':ts_month[i],'y':yr[i],'m':mk[i],'dir':'L','pnl':-loss,'r':'FL','bal':bal})
                in_pos=False
                if bal>peak:peak=bal
                dd2=(peak-bal)/peak if peak>0 else 0
                if dd2>mdd:mdd=dd2;continue
            if p_dir==-1 and h>=p_entry*(1+inv):
                loss=p_margin;bal-=loss
                trades.append({'t':ts_month[i],'y':yr[i],'m':mk[i],'dir':'S','pnl':-loss,'r':'FL','bal':bal})
                in_pos=False
                if bal>peak:peak=bal
                dd2=(peak-bal)/peak if peak>0 else 0
                if dd2>mdd:mdd=dd2;continue

            # SL check (skip if TSL active and tsl_disables_sl)
            sl_check=True
            if tsl_disables_sl and tsl_active:sl_check=False

            if sl_check:
                if p_dir==1 and l<=p_sl:
                    pnl=p_size*(p_sl-p_entry)/p_entry-p_size*fee;bal+=pnl
                    trades.append({'t':ts_month[i],'y':yr[i],'m':mk[i],'dir':'L','pnl':pnl,'r':'SL','bal':bal})
                    in_pos=False
                elif p_dir==-1 and h>=p_sl:
                    pnl=p_size*(p_entry-p_sl)/p_entry-p_size*fee;bal+=pnl
                    trades.append({'t':ts_month[i],'y':yr[i],'m':mk[i],'dir':'S','pnl':pnl,'r':'SL','bal':bal})
                    in_pos=False

            if in_pos:
                # TSL tracking
                if p_dir==1:
                    if h>p_high:p_high=h
                    # TSL activation check (on high/low for v32.x)
                    if (p_high-p_entry)/p_entry>=ta_pct:
                        tsl_active=True
                        ns=p_high*(1-tp_pct)
                        if ns>tsl_price:tsl_price=ns
                    # TSL trigger (on close for v32.x)
                    if tsl_active and c<=tsl_price:
                        pnl=p_size*(c-p_entry)/p_entry-p_size*fee;bal+=pnl
                        trades.append({'t':ts_month[i],'y':yr[i],'m':mk[i],'dir':'L','pnl':pnl,'r':'TSL','bal':bal})
                        in_pos=False
                else:
                    if l<p_low:p_low=l
                    if (p_entry-p_low)/p_entry>=ta_pct:
                        tsl_active=True
                        ns=p_low*(1+tp_pct)
                        if tsl_price==0 or ns<tsl_price:tsl_price=ns
                    if tsl_active and c>=tsl_price:
                        pnl=p_size*(p_entry-c)/p_entry-p_size*fee;bal+=pnl
                        trades.append({'t':ts_month[i],'y':yr[i],'m':mk[i],'dir':'S','pnl':pnl,'r':'TSL','bal':bal})
                        in_pos=False

        # --- SIGNAL DETECTION ---
        sig=0
        if i>=wu+1 and valid[i] and valid[i-1]:
            # Cross detection
            cross_up=above[i] and not above[i-1]
            cross_dn=not above[i] and above[i-1]

            if cross_up or cross_dn:
                cross_bar=i

            # For v32.x: monitor window approach
            if monitor_window>0:
                if i-cross_bar<=monitor_window and i>cross_bar:
                    # Check filters on current bar
                    adx_ok=adx_arr[i]>=adx_min if not np.isnan(adx_arr[i]) else False
                    rsi_ok=rsi_min<=rsi_arr[i]<=rsi_max if not np.isnan(rsi_arr[i]) else False
                    adx_rising=True
                    if adx_rise_bars>0 and i>=adx_rise_bars:
                        adx_rising=adx_arr[i]>adx_arr[i-adx_rise_bars] if not np.isnan(adx_arr[i-adx_rise_bars]) else False
                    gap_ok=True
                    if ema_gap_min>0 and not np.isnan(ma_fast[i]) and not np.isnan(ma_slow[i]):
                        gap=abs(ma_fast[i]-ma_slow[i])/ma_slow[i]
                        gap_ok=gap>=ema_gap_min

                    if adx_ok and rsi_ok and adx_rising and gap_ok:
                        if above[i]:sig=1  # Was bullish cross
                        else:sig=-1
                        cross_bar=-999  # Reset
            else:
                # Standard: immediate cross + filter
                adx_ok=adx_arr[i]>=adx_min if not np.isnan(adx_arr[i]) else False
                rsi_ok=rsi_min<=rsi_arr[i]<=rsi_max if not np.isnan(rsi_arr[i]) else False
                if strict:
                    if cross_up and adx_ok and rsi_ok and i>=2:sig=1
                    elif cross_dn and adx_ok and rsi_ok and i>=2:sig=-1
                else:
                    if cross_up and adx_ok and rsi_ok:sig=1
                    elif cross_dn and adx_ok and rsi_ok:sig=-1

        # Delay
        if sig!=0 and delay>0:pend=sig;pcnt=delay;sig=0
        if pcnt>0:
            pcnt-=1
            if pcnt==0:
                if pend==1 and valid[i] and ma_fast[i]>ma_slow[i]:sig=pend
                elif pend==-1 and valid[i] and ma_fast[i]<ma_slow[i]:sig=pend
                pend=0

        # Entry/exit
        if sig!=0:
            if in_pos:
                if p_dir!=sig:
                    pnl=p_size*(c-p_entry)/p_entry*p_dir-p_size*fee;bal+=pnl
                    trades.append({'t':ts_month[i],'y':yr[i],'m':mk[i],
                                  'dir':'L' if p_dir==1 else 'S','pnl':pnl,'r':'REV','bal':bal})
                    in_pos=False
                elif not skip_same:
                    pnl=p_size*(c-p_entry)/p_entry*p_dir-p_size*fee;bal+=pnl
                    trades.append({'t':ts_month[i],'y':yr[i],'m':mk[i],
                                  'dir':'L' if p_dir==1 else 'S','pnl':pnl,'r':'REV','bal':bal})
                    in_pos=False

            if not in_pos and not m_locked and bal>10:
                mg=bal*actual_margin;sz=mg*lev;bal-=sz*fee
                p_dir=sig;p_entry=c;p_size=sz;p_margin=mg
                p_sl=c*(1-sl_pct) if sig==1 else c*(1+sl_pct)
                p_high=c;p_low=c;tsl_active=False;tsl_price=0
                in_pos=True

        if bal>peak:peak=bal
        if peak>0:
            dd2=(peak-bal)/peak
            if dd2>mdd:mdd=dd2

    # Close remaining
    if in_pos:
        c=cl[-1];pnl=p_size*(c-p_entry)/p_entry*p_dir-p_size*fee;bal+=pnl
        trades.append({'t':ts_month[-1],'y':yr[-1],'m':mk[-1],
                      'dir':'L' if p_dir==1 else 'S','pnl':pnl,'r':'END','bal':bal})

    return compile_detail(trades,bal,init,mdd)


def compile_detail(trades,bal,init,mdd):
    tc=len(trades)
    if tc==0:
        return{'bal':round(bal,2),'ret':0.0,'pnl_usd':0.0,'tr':0,'pf':0.0,'mdd':0.0,
               'wr':0.0,'sl':0,'tsl':0,'rev':0,'fl':0,
               'yearly':{},'monthly':{}}

    pnls=np.array([t['pnl'] for t in trades])
    w=pnls>0;lo2=pnls<=0
    gp=pnls[w].sum() if w.any() else 0
    gl=abs(pnls[lo2].sum()) if lo2.any() else 0.001
    pf=min(gp/gl,999.99)

    # Yearly
    yearly={}
    for t in trades:
        y=int(t['y'])
        if y not in yearly:yearly[y]={'tr':0,'pnl':0.0,'w':0,'l':0,'sl':0,'tsl':0,'rev':0,'fl':0}
        yearly[y]['tr']+=1;yearly[y]['pnl']+=t['pnl']
        if t['pnl']>0:yearly[y]['w']+=1
        else:yearly[y]['l']+=1
        yearly[y][t['r'].lower()]=yearly[y].get(t['r'].lower(),0)+1
    for y in yearly:
        yp=sum(t['pnl'] for t in trades if int(t['y'])==y and t['pnl']>0)
        yl=abs(sum(t['pnl'] for t in trades if int(t['y'])==y and t['pnl']<=0))
        yearly[y]['pf']=round(min(yp/(yl if yl>0 else 0.001),999.99),2)
        yearly[y]['ret']=round(yearly[y]['pnl']/max(1,yearly[y].get('start_bal',init))*100,1)

    # Monthly
    monthly={};rb=init
    for t in trades:
        mk2='%d-%02d'%(t['m']//100,t['m']%100)
        if mk2 not in monthly:monthly[mk2]={'start':rb,'tr':0,'pnl':0.0,'sl':0,'tsl':0,'rev':0,'fl':0}
        monthly[mk2]['tr']+=1;monthly[mk2]['pnl']+=t['pnl']
        monthly[mk2][t['r'].lower()]=monthly[mk2].get(t['r'].lower(),0)+1
        monthly[mk2]['end']=t['bal'];rb=t['bal']
    for mk2 in monthly:
        s=monthly[mk2]['start'];e=monthly[mk2].get('end',s)
        monthly[mk2]['ret']=round((e-s)/s*100,2) if s>0 else 0

    return{
        'bal':round(bal,2),'ret':round((bal-init)/init*100,2),
        'pnl_usd':round(bal-init,2),
        'tr':tc,'pf':round(pf,2),'mdd':round(mdd*100,2),
        'wr':round(w.sum()/tc*100,2),
        'sl':sum(1 for t in trades if t['r']=='SL'),
        'tsl':sum(1 for t in trades if t['r']=='TSL'),
        'rev':sum(1 for t in trades if t['r']=='REV'),
        'fl':sum(1 for t in trades if t['r']=='FL'),
        'yearly':yearly,'monthly':monthly,
    }


# ============ STRATEGIES ============
STRATS=[
    {'f':'v23.4','mf':'EMA','fp':3,'ms':'EMA','sp':200,'ap':14,'at':35,'rp':14,'rm':30,'rx':65,
     'd':0,'sl':0.07,'ta':0.06,'tp':0.03,'l':10,'m':0.30,'init':5000,'ml':-0.20,
     'wu':250,'tsl_disables_sl':False,'monitor':0,'adx_rise':0,'ema_gap':0},

    {'f':'v32.2','mf':'EMA','fp':100,'ms':'EMA','sp':600,'ap':20,'at':30,'rp':10,'rm':40,'rx':80,
     'd':0,'sl':0.03,'ta':0.12,'tp':0.09,'l':10,'m':0.35,'init':5000,'ml':-0.20,
     'wu':650,'tsl_disables_sl':True,'monitor':24,'adx_rise':6,'ema_gap':0.002},

    {'f':'v32.3','mf':'EMA','fp':75,'ms':'SMA','sp':750,'ap':20,'at':30,'rp':11,'rm':40,'rx':80,
     'd':0,'sl':0.03,'ta':0.12,'tp':0.09,'l':10,'m':0.35,'init':5000,'ml':-0.20,
     'wu':800,'tsl_disables_sl':True,'monitor':24,'adx_rise':6,'ema_gap':0.002},

    {'f':'v15.4','mf':'EMA','fp':3,'ms':'EMA','sp':200,'ap':14,'at':35,'rp':14,'rm':30,'rx':65,
     'd':0,'sl':0.07,'ta':0.06,'tp':0.03,'l':10,'m':0.40,'init':5000,'ml':-0.30,
     'wu':250,'tsl_disables_sl':False,'monitor':0,'adx_rise':0,'ema_gap':0},

    {'f':'v15.5','mf':'EMA','fp':3,'ms':'EMA','sp':200,'ap':14,'at':35,'rp':14,'rm':35,'rx':65,
     'd':0,'sl':0.07,'ta':0.06,'tp':0.05,'l':10,'m':0.35,'init':5000,'ml':-0.25,
     'dd':-0.30,'dd_m':0.175,
     'wu':250,'tsl_disables_sl':False,'monitor':0,'adx_rise':0,'ema_gap':0},
]

ENGINES=[
    {'id':'E1','n':'Wilder','adx':'w','fee':0.0004,'slip':0,'strict':False,'skip':True},
    {'id':'E2','n':'EWM','adx':'e','fee':0.0004,'slip':0,'strict':False,'skip':True},
    {'id':'E3','n':'Strict','adx':'w','fee':0.0004,'slip':0,'strict':True,'skip':True},
    {'id':'E4','n':'Slip5bp','adx':'w','fee':0.0004,'slip':0.0005,'strict':False,'skip':True},
    {'id':'E5','n':'HiFee','adx':'w','fee':0.0006,'slip':0,'strict':False,'skip':True},
    {'id':'E6','n':'Reentry','adx':'w','fee':0.0004,'slip':0,'strict':False,'skip':False},
]

def main():
    t0=time.time()
    print('='*70)
    print('  기획서 검증 V1: 5 files x 6 engines = 30 backtests')
    print('='*70)

    # Load data
    print('Loading %s ...'%DATA)
    df=pd.read_csv(DATA,parse_dates=['timestamp']).sort_values('timestamp').reset_index(drop=True)
    d=df.set_index('timestamp')
    r30=d.resample('30min').agg({'open':'first','high':'max','low':'min','close':'last','volume':'sum'}).dropna().reset_index()
    ts=pd.to_datetime(r30['timestamp'].values)
    cl=r30['close'].values.astype(float);hi=r30['high'].values.astype(float)
    lo=r30['low'].values.astype(float)
    yr=ts.year.values.astype(np.int32);mk=(yr*100+ts.month.values).astype(np.int32)
    ts_month=mk
    n=len(cl)
    print('  30min: %d bars (%s ~ %s)'%(n,ts[0],ts[-1]))

    results=[]
    for s in STRATS:
        print('\n--- %s ---'%s['f'])
        # MA
        if s['mf']=='SMA':mf=_sma(cl,s['fp'])
        else:mf=_ema(cl,s['fp'])
        if s['ms']=='SMA':ms=_sma(cl,s['sp'])
        else:ms=_ema(cl,s['sp'])
        rsi_arr=calc_rsi(cl,s['rp'])

        row={'f':s['f'],'desc':'%s(%d)/%s(%d) L%dx M%d%%'%(s['mf'],s['fp'],s['ms'],s['sp'],s['l'],int(s['m']*100))}
        eng_data={}

        for eng in ENGINES:
            ax=adx_wilder(hi,lo,cl,s['ap']) if eng['adx']=='w' else adx_ewm(hi,lo,cl,s['ap'])
            if eng['slip']>0:
                np.random.seed(42)
                sl2=1+np.random.uniform(-eng['slip'],eng['slip'],n)
                cl2=cl*sl2;hi2=hi*sl2;lo2=lo*sl2
            else:cl2=cl;hi2=hi;lo2=lo

            r=backtest(cl2,hi2,lo2,yr,mk,ts_month,mf,ms,ax,rsi_arr,s,eng)
            eng_data[eng['id']]=r
            print('  %6s: $%12s (%+10.2f%%) Tr=%3d PF=%6.2f MDD=%5.2f%% SL=%d TSL=%d REV=%d FL=%d'%(
                eng['n'],'{:,.2f}'.format(r['bal']),r['ret'],r['tr'],r['pf'],r['mdd'],
                r['sl'],r['tsl'],r['rev'],r['fl']))

        rets=[eng_data[e['id']]['ret'] for e in ENGINES]
        pfs=[eng_data[e['id']]['pf'] for e in ENGINES]
        mdds=[eng_data[e['id']]['mdd'] for e in ENGINES]
        row['avg_ret']=round(np.mean(rets),2)
        row['min_ret']=round(np.min(rets),2)
        row['max_ret']=round(np.max(rets),2)
        row['avg_pf']=round(np.mean(pfs),2)
        row['min_pf']=round(np.min(pfs),2)
        row['avg_mdd']=round(np.mean(mdds),2)
        row['max_mdd']=round(np.max(mdds),2)
        row['avg_tr']=round(np.mean([eng_data[e['id']]['tr'] for e in ENGINES]),1)
        row['init']=s['init']
        row['engines']=eng_data
        results.append(row)

    # ============ RANKINGS ============
    print('\n'+'='*70)
    profitable=[r for r in results if r['min_ret']>0]
    by_ret=sorted(results,key=lambda x:x['avg_ret'],reverse=True)
    by_stable=sorted(results,key=lambda x:x['max_mdd'])
    by_recommend=sorted(profitable,key=lambda x:x['avg_ret']*x['avg_pf']/max(x['max_mdd'],1),reverse=True)
    discard=[r for r in results if r['min_ret']<=0 or r['min_pf']<1.0 or r['max_mdd']>60]

    print('\n수익률 TOP3:')
    for i,r in enumerate(by_ret[:3]):
        print('  %d. %s %+.2f%%'%(i+1,r['f'],r['avg_ret']))
    print('\n안정형 TOP3:')
    for i,r in enumerate(by_stable[:3]):
        print('  %d. %s MDD:%.2f%%'%(i+1,r['f'],r['max_mdd']))
    print('\n추천안 TOP3:')
    for i,r in enumerate(by_recommend[:3]):
        print('  %d. %s Score:%.1f'%(i+1,r['f'],r['avg_ret']*r['avg_pf']/max(r['max_mdd'],1)))
    print('\n폐기안:')
    for i,r in enumerate(discard[:3]):
        print('  %d. %s MinRet:%+.2f%% MaxMDD:%.2f%%'%(i+1,r['f'],r['min_ret'],r['max_mdd']))

    # ============ EXCEL ============
    print('\nGenerating Excel...')
    wb=Workbook()
    hdr_font=Font(bold=True,size=9,color='FFFFFF')
    hdr_fill=PatternFill('solid',fgColor='4472C4')
    green=PatternFill('solid',fgColor='C6EFCE')
    red=PatternFill('solid',fgColor='FFC7CE')
    yellow=PatternFill('solid',fgColor='FFEB9C')
    ctr=Alignment(horizontal='center',vertical='center',wrap_text=True)
    bdr=Border(left=Side('thin'),right=Side('thin'),top=Side('thin'),bottom=Side('thin'))

    def style_header(ws,row,ncols):
        for j in range(1,ncols+1):
            c=ws.cell(row=row,column=j)
            c.font=hdr_font;c.fill=hdr_fill;c.alignment=ctr;c.border=bdr

    def style_cell(ws,row,col,val,fmt=None):
        c=ws.cell(row=row,column=col,value=val)
        c.font=Font(size=8);c.alignment=ctr;c.border=bdr
        if fmt:c.number_format=fmt
        return c

    # Sheet 1: 종합 요약
    ws1=wb.active;ws1.title='종합 요약'
    h1=['순위','파일명','전략','평균수익률(%)','최소수익률(%)','평균PF','최소PF',
        '평균MDD(%)','최대MDD(%)','평균거래','판정']
    for e in ENGINES:h1+=['[%s]수익%%'%e['n'],'[%s]PF'%e['n'],'[%s]MDD%%'%e['n'],'[%s]거래'%e['n'],'[%s]SL'%e['n']]
    for j,h in enumerate(h1):ws1.cell(row=1,column=j+1,value=h)
    style_header(ws1,1,len(h1))

    for i,r in enumerate(sorted(results,key=lambda x:x['avg_ret'],reverse=True)):
        rn=i+2
        style_cell(ws1,rn,1,i+1)
        style_cell(ws1,rn,2,r['f'])
        style_cell(ws1,rn,3,r['desc'])
        style_cell(ws1,rn,4,r['avg_ret'],'#,##0.00')
        style_cell(ws1,rn,5,r['min_ret'],'#,##0.00')
        style_cell(ws1,rn,6,r['avg_pf'],'#,##0.00')
        style_cell(ws1,rn,7,r['min_pf'],'#,##0.00')
        style_cell(ws1,rn,8,r['avg_mdd'],'#,##0.00')
        style_cell(ws1,rn,9,r['max_mdd'],'#,##0.00')
        style_cell(ws1,rn,10,r['avg_tr'],'#,##0.0')
        judge='추천' if r['min_ret']>0 and r['max_mdd']<=50 else ('주의' if r['min_ret']>0 else '폐기')
        c=style_cell(ws1,rn,11,judge)
        c.fill=green if judge=='추천' else (yellow if judge=='주의' else red)

        col=12
        for e in ENGINES:
            er=r['engines'][e['id']]
            style_cell(ws1,rn,col,er['ret'],'#,##0.00');col+=1
            style_cell(ws1,rn,col,er['pf'],'#,##0.00');col+=1
            style_cell(ws1,rn,col,er['mdd'],'#,##0.00');col+=1
            style_cell(ws1,rn,col,er['tr']);col+=1
            style_cell(ws1,rn,col,er['sl']);col+=1

    ws1.column_dimensions['A'].width=5;ws1.column_dimensions['B'].width=12;ws1.column_dimensions['C'].width=28
    for j in range(4,col):ws1.column_dimensions[get_column_letter(j)].width=10

    # Sheet 2-6: 각 기획서별 월별 상세 (E1 Wilder 기준)
    for s_r in results:
        ws=wb.create_sheet(s_r['f'][:20])
        e1=s_r['engines']['E1']

        # Header info
        ws.cell(row=1,column=1,value='기획서: %s'%s_r['f']).font=Font(bold=True,size=11)
        ws.cell(row=2,column=1,value='전략: %s'%s_r['desc']).font=Font(size=9)
        ws.cell(row=3,column=1,value='잔액: $%s | 수익률: %+.2f%% | PF: %.2f | MDD: %.2f%% | 거래: %d'%(
            '{:,.2f}'.format(e1['bal']),e1['ret'],e1['pf'],e1['mdd'],e1['tr'])).font=Font(size=9)

        # Yearly table
        ws.cell(row=5,column=1,value='[연도별 성과]').font=Font(bold=True,size=10)
        yh=['연도','거래','승','패','PnL($)','PF','SL','TSL','REV','FL']
        for j,h in enumerate(yh):ws.cell(row=6,column=j+1,value=h)
        style_header(ws,6,len(yh))

        yrly=e1.get('yearly',{})
        row_n=7
        for y in sorted(yrly.keys()):
            yd=yrly[y]
            style_cell(ws,row_n,1,y)
            style_cell(ws,row_n,2,yd['tr'])
            style_cell(ws,row_n,3,yd.get('w',0))
            style_cell(ws,row_n,4,yd.get('l',0))
            style_cell(ws,row_n,5,round(yd['pnl'],2),'#,##0.00')
            style_cell(ws,row_n,6,yd.get('pf',0),'#,##0.00')
            style_cell(ws,row_n,7,yd.get('sl',0))
            style_cell(ws,row_n,8,yd.get('tsl',0))
            style_cell(ws,row_n,9,yd.get('rev',0))
            style_cell(ws,row_n,10,yd.get('fl',0))
            # Color PnL
            if yd['pnl']<0:ws.cell(row=row_n,column=5).fill=red
            else:ws.cell(row=row_n,column=5).fill=green
            row_n+=1

        # Monthly table
        row_n+=1
        ws.cell(row=row_n,column=1,value='[월별 상세]').font=Font(bold=True,size=10)
        row_n+=1
        mh=['월','수익률(%)','거래','PnL($)','잔액($)','SL','TSL','REV','FL']
        for j,h in enumerate(mh):ws.cell(row=row_n,column=j+1,value=h)
        style_header(ws,row_n,len(mh))

        monthly=e1.get('monthly',{})
        row_n+=1
        for mk2 in sorted(monthly.keys()):
            md=monthly[mk2]
            if md['tr']>0:
                style_cell(ws,row_n,1,mk2)
                style_cell(ws,row_n,2,md['ret'],'#,##0.00')
                style_cell(ws,row_n,3,md['tr'])
                style_cell(ws,row_n,4,round(md['pnl'],2),'#,##0.00')
                style_cell(ws,row_n,5,round(md.get('end',0),2),'#,##0.00')
                style_cell(ws,row_n,6,md.get('sl',0))
                style_cell(ws,row_n,7,md.get('tsl',0))
                style_cell(ws,row_n,8,md.get('rev',0))
                style_cell(ws,row_n,9,md.get('fl',0))
                if md['ret']<0:ws.cell(row=row_n,column=2).fill=red
                elif md['ret']>0:ws.cell(row=row_n,column=2).fill=green
                row_n+=1

        # 6엔진 비교
        row_n+=1
        ws.cell(row=row_n,column=1,value='[6엔진 교차검증]').font=Font(bold=True,size=10)
        row_n+=1
        eh=['엔진','수익률(%)','PF','MDD(%)','거래','SL','TSL','REV','FL','잔액($)']
        for j,h in enumerate(eh):ws.cell(row=row_n,column=j+1,value=h)
        style_header(ws,row_n,len(eh))
        row_n+=1
        for e in ENGINES:
            er=s_r['engines'][e['id']]
            style_cell(ws,row_n,1,e['n'])
            style_cell(ws,row_n,2,er['ret'],'#,##0.00')
            style_cell(ws,row_n,3,er['pf'],'#,##0.00')
            style_cell(ws,row_n,4,er['mdd'],'#,##0.00')
            style_cell(ws,row_n,5,er['tr'])
            style_cell(ws,row_n,6,er['sl'])
            style_cell(ws,row_n,7,er['tsl'])
            style_cell(ws,row_n,8,er['rev'])
            style_cell(ws,row_n,9,er['fl'])
            style_cell(ws,row_n,10,er['bal'],'#,##0.00')
            row_n+=1

        for j in range(1,11):ws.column_dimensions[get_column_letter(j)].width=12

    # Rankings sheet
    ws_rank=wb.create_sheet('순위 종합')
    rankings=[
        ('수익률 TOP3',sorted(results,key=lambda x:x['avg_ret'],reverse=True)[:3]),
        ('안정형 TOP3',sorted(results,key=lambda x:x['max_mdd'])[:3]),
        ('추천안 TOP3',sorted(profitable,key=lambda x:x['avg_ret']*x['avg_pf']/max(x['max_mdd'],1),reverse=True)[:3] if profitable else []),
        ('폐기안 TOP3',discard[:3]),
    ]
    row_n=1
    for title,ranked in rankings:
        ws_rank.cell(row=row_n,column=1,value=title).font=Font(bold=True,size=11)
        row_n+=1
        rh=['순위','파일명','평균수익률(%)','평균PF','최대MDD(%)','평균거래','사유']
        for j,h in enumerate(rh):ws_rank.cell(row=row_n,column=j+1,value=h)
        style_header(ws_rank,row_n,len(rh))
        row_n+=1
        for i,r in enumerate(ranked):
            style_cell(ws_rank,row_n,1,i+1)
            style_cell(ws_rank,row_n,2,r['f'])
            style_cell(ws_rank,row_n,3,r['avg_ret'],'#,##0.00')
            style_cell(ws_rank,row_n,4,r['avg_pf'],'#,##0.00')
            style_cell(ws_rank,row_n,5,r['max_mdd'],'#,##0.00')
            style_cell(ws_rank,row_n,6,r['avg_tr'],'#,##0.0')
            reasons=[]
            if 'TOP' in title and '수익' in title:reasons.append('최고수익')
            if '안정' in title:reasons.append('최저MDD')
            if '추천' in title:reasons.append('수익×PF/MDD 최적')
            if '폐기' in title:
                if r['min_ret']<=0:reasons.append('손실')
                if r['max_mdd']>60:reasons.append('MDD>60%')
                if r['min_pf']<1:reasons.append('PF<1')
            style_cell(ws_rank,row_n,7,', '.join(reasons))
            row_n+=1
        row_n+=1

    for j in range(1,8):ws_rank.column_dimensions[get_column_letter(j)].width=14

    wb.save(OUT)
    print('\nExcel saved: %s'%OUT)
    print('Sheets: 종합요약 + %d개 기획서별 월별상세 + 순위종합'%len(results))
    print('\nDone in %.1f min'%((time.time()-t0)/60))

if __name__=='__main__':
    main()
